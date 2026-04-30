from fastapi import FastAPI, APIRouter, UploadFile, File, Form, HTTPException, Query, Body, Depends, Header
from fastapi.responses import HTMLResponse, StreamingResponse
from fastapi.security import HTTPBearer, HTTPAuthorizationCredentials
from dotenv import load_dotenv
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
import os
import logging
import re
import uuid
import aiohttp
import math
import smtplib
import httpx
import io
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from pathlib import Path
from pydantic import BaseModel, Field, EmailStr
from typing import List, Optional
from datetime import datetime, timezone, timedelta
from bs4 import BeautifulSoup
import xml.etree.ElementTree as ET
from jose import JWTError, jwt
from passlib.context import CryptContext
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

mongo_url = os.environ['MONGO_URL']
client = AsyncIOMotorClient(mongo_url)
db = client[os.environ['DB_NAME']]

app = FastAPI()
api_router = APIRouter(prefix="/api")

logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(name)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

# ============ JWT & AUTH CONFIGURATION ============

JWT_SECRET_KEY = os.environ.get("JWT_SECRET_KEY", "apodixxi-secret-key-change-me")
JWT_ALGORITHM = os.environ.get("JWT_ALGORITHM", "HS256")
ACCESS_TOKEN_EXPIRE_MINUTES = int(os.environ.get("ACCESS_TOKEN_EXPIRE_MINUTES", "30"))
REFRESH_TOKEN_EXPIRE_DAYS = int(os.environ.get("REFRESH_TOKEN_EXPIRE_DAYS", "7"))

# Password hashing
pwd_context = CryptContext(schemes=["bcrypt"], deprecated="auto")
security = HTTPBearer(auto_error=False)

def hash_password(password: str) -> str:
    return pwd_context.hash(password)

def verify_password(plain_password: str, hashed_password: str) -> bool:
    try:
        return pwd_context.verify(plain_password, hashed_password)
    except Exception:
        return False

def create_access_token(user_id: str, email: str, expires_delta: Optional[timedelta] = None) -> str:
    if expires_delta:
        expire = datetime.utcnow() + expires_delta
    else:
        expire = datetime.utcnow() + timedelta(minutes=ACCESS_TOKEN_EXPIRE_MINUTES)
    
    to_encode = {
        "sub": user_id,
        "email": email,
        "exp": expire,
        "iat": datetime.utcnow(),
        "type": "access"
    }
    return jwt.encode(to_encode, JWT_SECRET_KEY, algorithm=JWT_ALGORITHM)

def create_refresh_token(user_id: str) -> str:
    expire = datetime.utcnow() + timedelta(days=REFRESH_TOKEN_EXPIRE_DAYS)
    to_encode = {
        "sub": user_id,
        "exp": expire,
        "type": "refresh"
    }
    return jwt.encode(to_encode, JWT_SECRET_KEY, algorithm=JWT_ALGORITHM)

def verify_token(token: str) -> dict:
    try:
        payload = jwt.decode(token, JWT_SECRET_KEY, algorithms=[JWT_ALGORITHM])
        return payload
    except JWTError as e:
        raise HTTPException(status_code=401, detail=f"Invalid token: {str(e)}")

async def get_current_user(credentials: HTTPAuthorizationCredentials = Depends(security)):
    if not credentials:
        raise HTTPException(status_code=401, detail="Not authenticated")
    
    payload = verify_token(credentials.credentials)
    user_id = payload.get("sub")
    if not user_id:
        raise HTTPException(status_code=401, detail="Invalid token")
    
    user = await db.users.find_one({"_id": user_id})
    if not user:
        raise HTTPException(status_code=401, detail="User not found")
    
    return user

async def get_optional_user(credentials: HTTPAuthorizationCredentials = Depends(security)):
    """Get current user if token provided, otherwise return None"""
    if not credentials:
        return None
    try:
        payload = verify_token(credentials.credentials)
        user_id = payload.get("sub")
        if user_id:
            return await db.users.find_one({"_id": user_id})
    except:
        pass
    return None

# ============ AUTH MODELS ============

class UserSignupRequest(BaseModel):
    email: str
    password: str
    name: str = ""

class UserLoginRequest(BaseModel):
    email: str
    password: str

class GoogleAuthRequest(BaseModel):
    google_id: Optional[str] = None
    id_token: Optional[str] = None
    email: str
    name: str = ""
    picture: Optional[str] = None

class AppleAuthRequest(BaseModel):
    apple_id: Optional[str] = None
    identity_token: Optional[str] = None
    email: str
    name: str = ""

class FacebookAuthRequest(BaseModel):
    facebook_id: Optional[str] = None
    access_token: Optional[str] = None
    email: str
    name: str = ""
    picture: Optional[str] = None
    post_to_wall: bool = False

class PhoneOTPRequest(BaseModel):
    phone_number: str

class PhoneOTPVerifyRequest(BaseModel):
    phone_number: str
    otp: str

class FirebasePhoneVerifyRequest(BaseModel):
    phone_number: str
    firebase_uid: str

class PhoneCompleteRequest(BaseModel):
    phone_number: str
    email: str

class UpdatePhoneRequest(BaseModel):
    phone_number: str

class TokenResponse(BaseModel):
    access_token: str
    refresh_token: str
    token_type: str = "bearer"
    user: dict

class PromoCodeRequest(BaseModel):
    code: str

def safe_float(value, default=0.0):
    """Convert value to float, handling inf/nan and invalid values."""
    try:
        f = float(value)
        if math.isnan(f) or math.isinf(f):
            return default
        return f
    except (ValueError, TypeError):
        return default

def sanitize_receipt_data(data):
    """Sanitize all float values in receipt data to prevent JSON serialization errors."""
    if isinstance(data, dict):
        return {k: sanitize_receipt_data(v) for k, v in data.items()}
    elif isinstance(data, list):
        return [sanitize_receipt_data(item) for item in data]
    elif isinstance(data, float):
        if math.isnan(data) or math.isinf(data):
            return 0.0
        return data
    return data

# ── Models ──

class ProductItem(BaseModel):
    code: str = ""
    description: str = ""
    unit: str = ""
    quantity: float = 1.0
    unit_price: float = 0.0
    pre_discount_value: float = 0.0
    discount: float = 0.0
    vat_percent: float = 0.0
    total_value: float = 0.0

class ReceiptData(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    device_id: str = ""
    store_name: str = ""
    store_address: str = ""
    store_vat: str = ""
    receipt_number: str = ""
    date: str = ""
    payment_method: str = ""
    items: List[ProductItem] = []
    subtotal: float = 0.0
    discount_total: float = 0.0
    total: float = 0.0
    vat_total: float = 0.0
    net_total: float = 0.0
    source_url: str = ""
    source_type: str = ""  # entersoft, impact, xml, manual
    provider: str = ""
    created_at: str = Field(default_factory=lambda: datetime.now(timezone.utc).isoformat())

class ManualReceiptInput(BaseModel):
    device_id: str
    store_name: str
    date: str
    items: List[ProductItem]
    total: float
    payment_method: str = ""

class URLImportInput(BaseModel):
    device_id: str
    url: str
    force_import: bool = False  # If true, import even if duplicate

class DeviceRegister(BaseModel):
    device_id: str
    language: str = "el"

class WebViewExtractedData(BaseModel):
    device_id: str
    url: str = ""
    raw_text: str = ""
    items: List[dict] = []
    store_name: str = ""
    found_final_total: float = 0.0  # Final total with VAT found in raw text

# ── Parsers ──

# VAT to Store Name mapping - Major Greek supermarket chains
STORE_VAT_MAPPING = {
    # 1. ΣΚΛΑΒΕΝΙΤΗΣ
    "800764388": "ΣΚΛΑΒΕΝΙΤΗΣ",
    
    # 2. ΑΒ ΒΑΣΙΛΟΠΟΥΛΟΣ
    "094025817": "ΑΒ ΒΑΣΙΛΟΠΟΥΛΟΣ",
    "094014249": "ΑΒ ΒΑΣΙΛΟΠΟΥΛΟΣ",  # Alternative VAT
    "094059506": "ΑΒ ΒΑΣΙΛΟΠΟΥΛΟΣ",  # Alternative VAT
    
    # 3. METRO
    "094062259": "METRO",
    
    # 4. ΜΑΣΟΥΤΗΣ
    "094063140": "ΜΑΣΟΥΤΗΣ",
    "094063169": "ΜΑΣΟΥΤΗΣ",  # Alternative VAT
    
    # 5. ΚΡΗΤΙΚΟΣ
    "094247924": "ΚΡΗΤΙΚΟΣ",
    
    # 6. ΓΑΛΑΞΙΑΣ (ΠΕΝΤΕ ΑΕ)
    "094116278": "ΓΑΛΑΞΙΑΣ",
    
    # 7. MARKET IN
    "998771189": "MARKET IN",
    "800469072": "MARKET IN",  # Alternative VAT
    
    # 8. BAZAAR
    "094384144": "BAZAAR",
    "094288618": "BAZAAR",  # Alternative VAT
    
    # 9. ΕΓΝΑΤΙΑ
    "094357707": "ΕΓΝΑΤΙΑ",
    
    # 10. ΣΥΝ.ΚΑ ΚΡΗΤΗΣ
    "996722071": "ΣΥΝ.ΚΑ ΚΡΗΤΗΣ",
    "096070396": "ΣΥΝ.ΚΑ ΚΡΗΤΗΣ",
    
    # Other known stores
    "800424460": "LIDL",
    "099326240": "JUMBO",
    "094281307": "MY MARKET",
    
    # THE MART
    "094021972": "THE MART",
    
    # ΜΟΥΣΤΑΚΑΣ (Toy Store)
    "094150585": "ΜΟΥΣΤΑΚΑΣ",
}

# Keywords to detect store brand from name (for franchises with different VAT)
STORE_BRAND_KEYWORDS = {
    "ΣΚΛΑΒΕΝΙΤ": "ΣΚΛΑΒΕΝΙΤΗΣ",
    "SKLAVENITIS": "ΣΚΛΑΒΕΝΙΤΗΣ",
    "ΒΑΣΙΛΟΠΟΥΛ": "ΑΒ ΒΑΣΙΛΟΠΟΥΛΟΣ",
    "ALFA VITA": "ΑΒ ΒΑΣΙΛΟΠΟΥΛΟΣ",
    "ΑΛΦΑ ΒΗΤΑ": "ΑΒ ΒΑΣΙΛΟΠΟΥΛΟΣ",
    "A.B.": "ΑΒ ΒΑΣΙΛΟΠΟΥΛΟΣ",
    "AB ": "ΑΒ ΒΑΣΙΛΟΠΟΥΛΟΣ",
    "METRO": "METRO",
    "ΜΕΤΡΟ": "METRO",
    "ΜΑΣΟΥΤ": "ΜΑΣΟΥΤΗΣ",
    "MASOUTIS": "ΜΑΣΟΥΤΗΣ",
    "ΚΡΗΤΙΚΟ": "ΚΡΗΤΙΚΟΣ",
    "KRITIKOS": "ΚΡΗΤΙΚΟΣ",
    "ΓΑΛΑΞΙΑ": "ΓΑΛΑΞΙΑΣ",
    "GALAXIAS": "ΓΑΛΑΞΙΑΣ",
    "ΠΕΝΤΕ": "ΓΑΛΑΞΙΑΣ",
    "MARKET IN": "MARKET IN",
    "MARKETIN": "MARKET IN",
    "BAZAAR": "BAZAAR",
    "ΜΠΑΖΑΡ": "BAZAAR",
    "ΕΓΝΑΤΙΑ": "ΕΓΝΑΤΙΑ",
    "EGNATIA": "ΕΓΝΑΤΙΑ",
    "ΣΥΝ.ΚΑ": "ΣΥΝ.ΚΑ ΚΡΗΤΗΣ",
    "ΣΥΝΚΑ": "ΣΥΝ.ΚΑ ΚΡΗΤΗΣ",
    "LIDL": "LIDL",
    "ΛΙΝΤΛ": "LIDL",
    "JUMBO": "JUMBO",
    "ΤΖΑΜΠΟ": "JUMBO",
    "MY MARKET": "MY MARKET",
    "MYMARKET": "MY MARKET",
    "THE MART": "THE MART",
    "THEMART": "THE MART",
    "ΜΟΥΣΤΑΚΑΣ": "ΜΟΥΣΤΑΚΑΣ",
    "MOUSTAKAS": "ΜΟΥΣΤΑΚΑΣ",
}

def get_store_name_from_vat(vat: str, fallback: str = "") -> str:
    """Get clean store name from VAT number."""
    if vat and vat in STORE_VAT_MAPPING:
        return STORE_VAT_MAPPING[vat]
    return fallback

def detect_store_brand(store_name: str) -> str:
    """Detect store brand from name using keywords (for franchises)."""
    if not store_name:
        return ""
    name_upper = store_name.upper()
    for keyword, brand in STORE_BRAND_KEYWORDS.items():
        if keyword in name_upper:
            return brand
    return ""

def get_clean_store_name(vat: str, raw_name: str) -> str:
    """
    Get the cleanest store name possible:
    1. First try VAT mapping (most reliable)
    2. Then try keyword detection from name (for franchises)
    3. Fallback to raw name
    """
    # Try VAT mapping first
    if vat:
        mapped = get_store_name_from_vat(vat)
        if mapped:
            return mapped
    
    # Try keyword detection for franchises
    if raw_name:
        brand = detect_store_brand(raw_name)
        if brand:
            return brand
    
    # Fallback to raw name (cleaned up)
    if raw_name:
        # Remove common suffixes like ΑΕ, ΑΕΒΕ, etc.
        cleaned = raw_name.strip()
        for suffix in [" ΑΝΩΝΥΜΗ ΕΤΑΙΡΕΙΑ", " ΜΟΝΟΠΡΟΣΩΠΗ", " ΑΕΒΕ", " Α.Ε.", " ΑΕ", " ΕΠΕ", " ΙΚΕ"]:
            if cleaned.upper().endswith(suffix):
                cleaned = cleaned[:-len(suffix)].strip()
        return cleaned
    
    return "Άγνωστο Κατάστημα"

def parse_greek_number(text: str) -> float:
    """Parse Greek-formatted numbers (uses comma as decimal separator)."""
    if not text:
        return 0.0
    text = text.strip().replace('\xa0', '').replace('EUR', '').replace('€', '').strip()
    
    # Handle different number formats:
    # Greek: 1.234,56 or 1234,56
    # English: 1,234.56 or 1234.56
    
    # Count dots and commas
    dots = text.count('.')
    commas = text.count(',')
    
    if commas == 1 and dots == 0:
        # Format: 1234,56 - comma is decimal
        text = text.replace(',', '.')
    elif dots == 1 and commas == 0:
        # Format: 1234.56 - already correct
        pass
    elif dots >= 1 and commas == 1:
        # Format: 1.234,56 - dot is thousand separator, comma is decimal
        text = text.replace('.', '').replace(',', '.')
    elif commas >= 1 and dots == 1:
        # Format: 1,234.56 - comma is thousand separator, dot is decimal
        text = text.replace(',', '')
    else:
        # Just try to parse, removing non-numeric except decimal point
        text = text.replace('.', '').replace(',', '.')
    
    try:
        return float(text)
    except (ValueError, TypeError):
        return 0.0

async def fetch_html(url: str) -> str:
    headers = {
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36',
        'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8',
        'Accept-Language': 'el-GR,el;q=0.9,en;q=0.8',
    }
    async with aiohttp.ClientSession(headers=headers) as session:
        async with session.get(url, timeout=aiohttp.ClientTimeout(total=60), ssl=False) as resp:
            if resp.status != 200:
                raise HTTPException(status_code=400, detail=f"Could not fetch URL: HTTP {resp.status}")
            return await resp.text()

async def fetch_entersoft_invoice(url: str) -> str:
    """Fetch Entersoft invoice by first getting the page, extracting iframe src, then fetching iframe content."""
    html = await fetch_html(url)
    soup = BeautifulSoup(html, 'lxml')
    iframe = soup.find('iframe', id='iframeContent')
    if not iframe or not iframe.get('src'):
        raise HTTPException(status_code=400, detail="Could not find invoice iframe in page")
    iframe_src = iframe['src']
    if iframe_src.startswith('/'):
        iframe_src = 'https://e-invoicing.gr' + iframe_src
    return await fetch_html(iframe_src)

def parse_entersoft(html: str, source_url: str) -> dict:
    soup = BeautifulSoup(html, 'lxml')
    data = {
        "store_name": "",
        "store_address": "",
        "store_vat": "",
        "receipt_number": "",
        "date": "",
        "payment_method": "",
        "items": [],
        "subtotal": 0.0,
        "discount_total": 0.0,
        "total": 0.0,
        "vat_total": 0.0,
        "net_total": 0.0,
        "source_url": source_url,
        "source_type": "entersoft",
        "provider": "Entersoft"
    }
    header = soup.find('div', class_='BoldBlueHeader fontSize12pt')
    if header:
        data["store_name"] = header.get_text(strip=True)

    for div in soup.find_all('div', class_='fontSize8pt'):
        txt = div.get_text(strip=True)
        if 'Α.Φ.Μ' in txt:
            match = re.search(r'Α\.Φ\.Μ:\s*(\d+)', txt)
            if match:
                data["store_vat"] = match.group(1)
            if ',' in txt and 'Α.Φ.Μ' in txt:
                addr_part = txt.split('Α.Φ.Μ')[0].strip().rstrip(',')
                if not data["store_address"]:
                    data["store_address"] = addr_part
        elif txt.startswith('Αρ. Παραστατικού'):
            data["receipt_number"] = txt.replace('Αρ. Παραστατικού:', '').strip()
        elif txt.startswith('Ημ/νία έκδοσης'):
            data["date"] = txt.replace('Ημ/νία έκδοσης:', '').strip()

    # Try address from a specific div
    addr_divs = soup.find_all('div', class_='fontSize8pt')
    for d in addr_divs:
        t = d.get_text(strip=True)
        if t and not t.startswith('Α.Φ.Μ') and not t.startswith('Αρ.') and not t.startswith('Ημ/') and not t.startswith('Email') and not t.startswith('Πάροχος') and not t.startswith('Url') and not t.startswith('Άδεια') and not t.startswith('UID') and not t.startswith('M.AR') and not t.startswith('Auth') and t and ',' in t and any(c.isdigit() for c in t):
            if not data["store_address"] or len(t) < len(data["store_address"]):
                data["store_address"] = t
                break

    # Payment
    for div in soup.find_all('div', class_='fontSize8pt'):
        txt = div.get_text(strip=True)
        if txt in ['POS / e-POS', 'Μετρητά', 'Κάρτα']:
            data["payment_method"] = txt
            break

    table = soup.find('table', class_='table')
    if table:
        rows = table.find('tbody')
        if rows:
            for row in rows.find_all('tr'):
                cells = row.find_all('td')
                if len(cells) >= 9:
                    item = {
                        "code": cells[0].get_text(strip=True),
                        "description": cells[1].get_text(strip=True),
                        "unit": cells[2].get_text(strip=True),
                        "quantity": parse_greek_number(cells[3].get_text(strip=True)),
                        "unit_price": parse_greek_number(cells[4].get_text(strip=True)),
                        "pre_discount_value": parse_greek_number(cells[5].get_text(strip=True)),
                        "discount": parse_greek_number(cells[6].get_text(strip=True)),
                        "vat_percent": parse_greek_number(cells[7].get_text(strip=True).replace('%', '')),
                        "total_value": parse_greek_number(cells[8].get_text(strip=True)),
                    }
                    data["items"].append(item)

    # Totals
    for div in soup.find_all('div', class_='fontSize8pt'):
        txt = div.get_text(strip=True)
        if 'Αξία προ έκπτωσης' in txt:
            sibling = div.find_next_sibling('div')
            if sibling:
                data["subtotal"] = parse_greek_number(sibling.get_text(strip=True))
        elif txt.strip() == 'Έκπτωση':
            sibling = div.find_next_sibling('div')
            if sibling:
                data["discount_total"] = parse_greek_number(sibling.get_text(strip=True))

    for div in soup.find_all('div', class_='backgrey'):
        txt = div.get_text(strip=True)
        if 'EUR' in txt and any(c.isdigit() for c in txt):
            val = parse_greek_number(txt)
            if val > 0:
                data["total"] = val
                break

    # Try ΤΕΛΙΚΗ ΑΞΙΑ
    for div in soup.find_all('div'):
        txt = div.get_text(strip=True)
        if 'ΤΕΛΙΚΗ ΑΞΙΑ' in txt or 'ΤΕΛΙΚΗ' in txt:
            next_div = div.find_next_sibling('div')
            if next_div:
                val = parse_greek_number(next_div.get_text(strip=True))
                if val > 0:
                    data["total"] = val

    for div in soup.find_all('div', class_='fontSize6pt'):
        txt = div.get_text(strip=True)
        if 'Καθαρή Αξία' in txt:
            sibling = div.find_next_sibling('div')
            if sibling:
                data["net_total"] = parse_greek_number(sibling.get_text(strip=True))
        elif 'Φ.Π.Α' in txt and 'ΑΝΑΛΥΣΗ' not in txt:
            sibling = div.find_next_sibling('div')
            if sibling:
                data["vat_total"] = parse_greek_number(sibling.get_text(strip=True))

    if not data["total"] and data["items"]:
        data["total"] = sum(i["total_value"] for i in data["items"])

    # Use clean store name (VAT mapping or keyword detection)
    data["store_name"] = get_clean_store_name(data["store_vat"], data["store_name"])

    return data


def parse_impact(html: str, source_url: str) -> dict:
    soup = BeautifulSoup(html, 'lxml')
    data = {
        "store_name": "",
        "store_address": "",
        "store_vat": "",
        "receipt_number": "",
        "date": "",
        "payment_method": "",
        "items": [],
        "subtotal": 0.0,
        "discount_total": 0.0,
        "total": 0.0,
        "vat_total": 0.0,
        "net_total": 0.0,
        "source_url": source_url,
        "source_type": "impact",
        "provider": "Impact/Entersoft One"
    }

    # Store name from issuer
    for span in soup.find_all('span', class_='value'):
        parent = span.find_parent('span', class_='field')
        if parent and 'field-IssuerName' in parent.get('class', []):
            data["store_name"] = span.get_text(strip=True)
        elif parent and 'field-IssuerVATNumber' in parent.get('class', []):
            data["store_vat"] = span.get_text(strip=True)

    # If no issuer name, try other patterns
    if not data["store_name"]:
        for span in soup.find_all('span', class_='field'):
            classes = span.get('class', [])
            if 'field-RegisteredName' in classes:
                val = span.find('span', class_='value')
                if val:
                    name = val.get_text(strip=True)
                    if name and name != 'Πελάτης Λιανικής':
                        data["store_name"] = name

    # Receipt number
    for span in soup.find_all('span', class_='field'):
        classes = span.get('class', [])
        if 'field-IssuerFormatedInvoiceSeriesNumber' in classes:
            val = span.find('span', class_='value')
            if val:
                data["receipt_number"] = val.get_text(strip=True)
        elif 'field-DateIssued' in classes:
            val = span.find('span', class_='value')
            if val:
                data["date"] = val.get_text(strip=True)

    # Items from table
    table = soup.find('table', class_='table')
    if table:
        tbody = table.find('tbody')
        if tbody:
            for row in tbody.find_all('tr'):
                cells = row.find_all('td')
                if len(cells) >= 7:
                    code_val = ""
                    desc_val = ""
                    qty_val = 0.0
                    unit_val = ""
                    price_val = 0.0
                    discount_val = 0.0
                    net_val = 0.0
                    vat_cat = 0.0
                    vat_total = 0.0
                    total_val = 0.0

                    for cell in cells:
                        header = cell.get('data-header', '')
                        val_span = cell.find('span', class_='value')
                        val_text = val_span.get_text(strip=True) if val_span else cell.get_text(strip=True)

                        if 'Κωδικός' in header:
                            code_val = val_text
                        elif 'Περιγραφή' in header:
                            desc_val = val_text
                        elif 'Ποσότητα' in header:
                            qty_val = parse_greek_number(val_text)
                        elif 'M.M.' in header or 'Μ.Μ.' in header:
                            unit_val = val_text
                        elif 'Τιμή' in header:
                            price_val = parse_greek_number(val_text)
                        elif 'Έκπτ' in header:
                            discount_val = parse_greek_number(val_text)
                        elif 'Καθαρή' in header:
                            net_val = parse_greek_number(val_text)
                        elif 'Κατηγορία' in header and 'Φ.Π.Α' in header:
                            vat_cat = parse_greek_number(val_text)
                        elif 'Σύνολο Φ.Π.Α' in header:
                            vat_total = parse_greek_number(val_text)
                        elif 'Τελικό' in header:
                            total_val = parse_greek_number(val_text)

                    if desc_val:
                        vat_pct = {1: 24, 2: 13, 3: 6, 4: 17, 5: 9}.get(int(vat_cat), 0) if vat_cat else 0
                        item = {
                            "code": code_val,
                            "description": desc_val,
                            "unit": unit_val,
                            "quantity": qty_val if qty_val else 1.0,
                            "unit_price": price_val,
                            "pre_discount_value": net_val + vat_total if net_val else total_val,
                            "discount": discount_val,
                            "vat_percent": float(vat_pct),
                            "total_value": total_val,
                        }
                        data["items"].append(item)

    if data["items"]:
        data["total"] = sum(i["total_value"] for i in data["items"])
        # Calculate net_total properly (don't multiply by 100)
        net_sum = 0.0
        for i in data["items"]:
            pre_disc = i.get("pre_discount_value", 0) or 0
            disc = i.get("discount", 0) or 0
            if isinstance(pre_disc, str):
                pre_disc = parse_greek_number(pre_disc)
            if isinstance(disc, str):
                disc = parse_greek_number(disc)
            net_sum += (pre_disc - disc)
        data["net_total"] = round(net_sum, 2)

    # Use clean store name (VAT mapping or keyword detection)
    data["store_name"] = get_clean_store_name(data["store_vat"], data["store_name"])

    return data


def parse_mydata_xml(xml_content: str) -> dict:
    data = {
        "store_name": "",
        "store_address": "",
        "store_vat": "",
        "receipt_number": "",
        "date": "",
        "payment_method": "",
        "items": [],
        "subtotal": 0.0,
        "discount_total": 0.0,
        "total": 0.0,
        "vat_total": 0.0,
        "net_total": 0.0,
        "source_url": "",
        "source_type": "xml",
        "provider": "Epsilon Digital (myData XML)"
    }

    try:
        root = ET.fromstring(xml_content)
    except ET.ParseError:
        raise HTTPException(status_code=400, detail="Invalid XML format")

    ns = {}
    for event, elem in ET.iterparse(__import__('io').StringIO(xml_content), events=['start-ns']):
        ns[elem[0]] = elem[1]

    def find_text(element, paths):
        for path in paths:
            found = element.find(path, ns) if ns else element.find(path)
            if found is not None and found.text:
                return found.text.strip()
        return ""

    # Try common myData XML structures
    issuer = root.find('.//{*}issuer') or root.find('.//issuer')
    if issuer is not None:
        data["store_vat"] = find_text(issuer, ['{*}vatNumber', 'vatNumber', '{*}afm', 'afm'])
        data["store_name"] = find_text(issuer, ['{*}name', 'name', '{*}registeredName', 'registeredName'])

    header = root.find('.//{*}invoiceHeader') or root.find('.//invoiceHeader')
    if header is not None:
        data["receipt_number"] = find_text(header, ['{*}invoiceSerialNumber', 'invoiceSerialNumber', '{*}aa', 'aa'])
        data["date"] = find_text(header, ['{*}issueDate', 'issueDate'])

    for detail in root.findall('.//{*}invoiceDetails') or root.findall('.//invoiceDetails') or root.findall('.//{*}InvoiceLine') or root.findall('.//InvoiceLine'):
        desc = find_text(detail, ['{*}itemDescr', 'itemDescr', '{*}description', 'description', '{*}name', 'name'])
        if not desc:
            continue
        qty_text = find_text(detail, ['{*}quantity', 'quantity'])
        net_text = find_text(detail, ['{*}netValue', 'netValue'])
        vat_text = find_text(detail, ['{*}vatAmount', 'vatAmount'])
        code_text = find_text(detail, ['{*}itemCode', 'itemCode', '{*}code', 'code'])

        qty = float(qty_text.replace(',', '.')) if qty_text else 1.0
        net_val = float(net_text.replace(',', '.')) if net_text else 0.0
        vat_val = float(vat_text.replace(',', '.')) if vat_text else 0.0
        total = net_val + vat_val

        item = {
            "code": code_text,
            "description": desc,
            "unit": find_text(detail, ['{*}measurementUnit', 'measurementUnit']) or "ΤΕΜ",
            "quantity": qty,
            "unit_price": round(net_val / qty, 5) if qty else net_val,
            "pre_discount_value": net_val,
            "discount": 0.0,
            "vat_percent": round((vat_val / net_val) * 100, 0) if net_val else 0.0,
            "total_value": total,
        }
        data["items"].append(item)

    if data["items"]:
        data["total"] = round(sum(i["total_value"] for i in data["items"]), 2)
        data["net_total"] = round(sum(i["pre_discount_value"] for i in data["items"]), 2)
        data["vat_total"] = round(data["total"] - data["net_total"], 2)

    summary = root.find('.//{*}invoiceSummary') or root.find('.//invoiceSummary')
    if summary is not None:
        t = find_text(summary, ['{*}totalGrossValue', 'totalGrossValue'])
        if t:
            data["total"] = float(t.replace(',', '.'))
        n = find_text(summary, ['{*}totalNetValue', 'totalNetValue'])
        if n:
            data["net_total"] = float(n.replace(',', '.'))
        v = find_text(summary, ['{*}totalVatAmount', 'totalVatAmount'])
        if v:
            data["vat_total"] = float(v.replace(',', '.'))

    return data


def detect_provider(url: str) -> str:
    if 'e-invoicing.gr' in url:
        return 'entersoft'
    elif 'einvoice.impact.gr' in url:
        return 'impact'
    elif 'epsilondigital' in url or 'epsilonnet.gr' in url:
        return 'epsilon_digital'
    return 'unknown'


def parse_webview_extracted(raw_text: str, items_from_dom: list, store_hint: str, source_url: str, found_final_total: float = 0.0) -> dict:
    """Parse data extracted from WebView DOM injection (Epsilon Digital pages)."""
    
    # Determine store name from URL if not provided
    detected_store = store_hint or ""
    if source_url:
        url_lower = source_url.lower()
        if 'marketin' in url_lower or 'market-in' in url_lower:
            detected_store = "MARKET IN"
        elif 'abmarket' in url_lower or 'epsilondigital-ab' in url_lower:
            detected_store = "ΑΒ ΒΑΣΙΛΟΠΟΥΛΟΣ"
        elif 'bazaar' in url_lower:
            detected_store = "BAZAAR"
        elif 'sklavenitis' in url_lower:
            detected_store = "ΣΚΛΑΒΕΝΙΤΗΣ"
    
    data = {
        "store_name": detected_store,
        "store_address": "",
        "store_vat": "",
        "receipt_number": "",
        "date": "",
        "payment_method": "",
        "items": [],
        "subtotal": 0.0,
        "discount_total": 0.0,
        "total": 0.0,
        "vat_total": 0.0,
        "net_total": 0.0,
        "source_url": source_url,
        "source_type": "webview",
        "provider": "Epsilon Digital (WebView)"
    }

    lines = raw_text.split('\n') if raw_text else []

    # Extract store info from raw text
    for line in lines:
        line = line.strip()
        if not line:
            continue
        # VAT
        afm_match = re.search(r'(?:Α\.?Φ\.?Μ\.?|ΑΦΜ)[:\s]*(\d{9})', line)
        if afm_match:
            data["store_vat"] = afm_match.group(1)
        # Date
        date_match = re.search(r'(\d{1,2}[/\-\.]\d{1,2}[/\-\.]\d{2,4})', line)
        if date_match and not data["date"]:
            data["date"] = date_match.group(1)
        # Receipt number
        if any(k in line for k in ['Αρ. Παραστατ', 'Αριθμός', 'Αρ. Τιμολ', 'Receipt']):
            num_match = re.search(r'[\d]+', line.split(':')[-1] if ':' in line else line)
            if num_match:
                data["receipt_number"] = num_match.group(0)
        # Payment
        if any(k in line for k in ['POS', 'Μετρητ', 'Κάρτα', 'ΠΛΗΡΩΜ']):
            data["payment_method"] = line.strip()

    # Use clean store name (VAT mapping or keyword detection)
    data["store_name"] = get_clean_store_name(data["store_vat"], data["store_name"])

    # Process items from DOM extraction (structured data from JS)
    # Helper function to check if this is a total row
    def is_total_row(text):
        if not text:
            return False
        t = text.upper().strip()
        return t in ['ΣΥΝΟΛΟ', 'ΣΥΝΟΛΑ', 'ΤΕΛΙΚΟ', 'TOTAL', 'ΠΛΗΡΩΤΕΟ'] or \
               (len(t) < 8 and re.match(r'^\d+[,.]?\d*$', t))

    # Helper function to check if this is a payment method row
    def is_payment_row(text):
        if not text:
            return False
        t = text.upper().strip()
        payment_keywords = ['EFT-POS', 'EFT POS', 'EFTPOS', 'POS', 'ΜΕΤΡΗΤΑ', 'ΚΑΡΤΑ',
                          'CASH', 'CARD', 'ΠΛΗΡΩΜ', 'VISA', 'MASTERCARD', 'CREDIT',
                          'DEBIT', 'PAYMENT', 'ΑΠΟΔ', 'ΡΕΣΤΑ', 'CHANGE']
        return any(kw in t for kw in payment_keywords)

    if items_from_dom:
        for item_raw in items_from_dom:
            code = str(item_raw.get('code', '')).strip()
            desc = str(item_raw.get('description', '')).strip()
            if not desc:
                continue
            
            # Skip total rows and payment rows
            if is_total_row(desc) or is_total_row(code):
                continue
            if is_payment_row(desc) or is_payment_row(code):
                continue

            qty_str = str(item_raw.get('quantity', '1')).replace(',', '.')
            price_str = str(item_raw.get('unit_price', item_raw.get('total', '0'))).replace(',', '.').replace('€', '')
            total_str = str(item_raw.get('total', '0')).replace(',', '.').replace('€', '')

            try:
                qty = float(qty_str) if qty_str else 1.0
            except ValueError:
                qty = 1.0
            try:
                total_val = float(total_str) if total_str else 0.0
            except ValueError:
                total_val = 0.0
            try:
                unit_price = float(price_str) if price_str else (total_val / qty if qty else total_val)
            except ValueError:
                unit_price = total_val / qty if qty else total_val

            item = {
                "code": code,
                "description": desc,
                "unit": str(item_raw.get('unit', 'ΤΕΜ')).strip() or 'ΤΕΜ',
                "quantity": qty,
                "unit_price": round(unit_price, 5),
                "pre_discount_value": round(total_val, 2),
                "discount": 0.0,
                "vat_percent": 0.0,
                "total_value": round(total_val, 2),
            }
            data["items"].append(item)

    # If no items from DOM, try parsing raw text for tab/space-separated product lines
    if not data["items"] and raw_text:
        for line in lines:
            parts = re.split(r'\t+', line.strip())
            if len(parts) >= 4:
                # Try: code, description, qty, price pattern
                possible_code = parts[0].strip()
                if possible_code and re.match(r'^\d+$', possible_code):
                    desc = parts[1].strip()
                    try:
                        total_val = float(parts[-1].replace(',', '.').replace('€', ''))
                    except ValueError:
                        continue
                    item = {
                        "code": possible_code,
                        "description": desc,
                        "unit": "ΤΕΜ",
                        "quantity": 1.0,
                        "unit_price": total_val,
                        "pre_discount_value": total_val,
                        "discount": 0.0,
                        "vat_percent": 0.0,
                        "total_value": total_val,
                    }
                    data["items"].append(item)

    if data["items"]:
        calculated_total = round(sum(i["total_value"] for i in data["items"]), 2)
        data["net_total"] = calculated_total
        
        # Use found_final_total only if it's provided and reasonably close to calculated
        # (within 20% tolerance for VAT differences)
        if found_final_total > 0:
            # If found_final_total is close to calculated, use calculated (more accurate)
            diff = abs(found_final_total - calculated_total)
            if diff <= calculated_total * 0.25:  # 25% tolerance
                data["total"] = calculated_total
            elif found_final_total >= calculated_total:
                # Only use found_final_total if it's larger (includes VAT we didn't capture)
                data["total"] = found_final_total
            else:
                data["total"] = calculated_total
        else:
            data["total"] = calculated_total

    # For Epsilon Digital, trust the calculated total from items more than raw text parsing
    # because raw text may contain confusing numbers (e.g., VAT totals, payment amounts)
    # Skip the additional total parsing from raw text

    return data


# ── API Routes ──

@api_router.get("/")
async def root():
    return {"message": "apodixxi API", "version": "1.0.0"}


# ============ AUTHENTICATION ENDPOINTS ============

@api_router.post("/auth/signup")
async def signup(request: UserSignupRequest):
    """Register a new user with email and password."""
    # Check if email already exists
    existing = await db.users.find_one({"email": request.email.lower()})
    if existing:
        raise HTTPException(status_code=400, detail="Email already registered")
    
    # Validate password
    if len(request.password) < 8:
        raise HTTPException(status_code=400, detail="Password must be at least 8 characters")
    
    # Create user
    user_id = str(uuid.uuid4())
    user = {
        "_id": user_id,
        "email": request.email.lower(),
        "name": request.name,
        "password_hash": hash_password(request.password),
        "phone": None,
        "auth_provider": "email",
        "account_type": "free",  # free or paid
        "subscription_expires_at": None,
        "is_email_verified": False,
        "created_at": datetime.now(timezone.utc).isoformat(),
        "last_login": datetime.now(timezone.utc).isoformat()
    }
    
    await db.users.insert_one(user)
    
    # Generate tokens
    access_token = create_access_token(user_id, user["email"])
    refresh_token = create_refresh_token(user_id)
    
    return {
        "access_token": access_token,
        "refresh_token": refresh_token,
        "token_type": "bearer",
        "user": {
            "id": user_id,
            "email": user["email"],
            "name": user["name"],
            "phone": user["phone"],
            "auth_provider": user["auth_provider"],
            "account_type": user["account_type"],
            "is_email_verified": user["is_email_verified"]
        }
    }


@api_router.post("/auth/login")
async def login(request: UserLoginRequest):
    """Login with email and password."""
    user = await db.users.find_one({"email": request.email.lower()})
    
    if not user:
        raise HTTPException(status_code=401, detail="Invalid email or password")
    
    if not user.get("password_hash"):
        raise HTTPException(status_code=401, detail="Please use your social login method")
    
    if not verify_password(request.password, user["password_hash"]):
        raise HTTPException(status_code=401, detail="Invalid email or password")
    
    # Get or create device_id for user
    user_device_id = user.get("device_id")
    if not user_device_id:
        user_device_id = f"dev_{uuid.uuid4().hex[:20]}"
    
    # Update last login and device_id
    await db.users.update_one(
        {"_id": user["_id"]},
        {"$set": {
            "last_login": datetime.now(timezone.utc).isoformat(),
            "device_id": user_device_id
        }}
    )
    
    # Generate tokens
    access_token = create_access_token(user["_id"], user["email"])
    refresh_token = create_refresh_token(user["_id"])
    
    return {
        "access_token": access_token,
        "refresh_token": refresh_token,
        "token_type": "bearer",
        "device_id": user_device_id,
        "user": {
            "id": user["_id"],
            "email": user["email"],
            "name": user.get("name", ""),
            "phone": user.get("phone"),
            "auth_provider": user.get("auth_provider", "email"),
            "account_type": user.get("account_type", "free"),
            "is_email_verified": user.get("is_email_verified", False),
            "device_id": user_device_id
        }
    }


@api_router.post("/auth/google")
async def google_auth(request: GoogleAuthRequest):
    """Authenticate with Google."""
    google_email = request.email.lower()
    
    # If google_id is provided directly (from mobile app), use it
    if request.google_id:
        # The user is authenticated via Google OAuth on mobile
        # We trust the data since it came from Google's API
        google_data = {
            "sub": request.google_id,
            "email": google_email,
            "name": request.name,
            "picture": request.picture
        }
    elif request.id_token:
        # Verify Google ID token (for web flow)
        try:
            async with httpx.AsyncClient() as client:
                response = await client.get(
                    "https://oauth2.googleapis.com/tokeninfo",
                    params={"id_token": request.id_token}
                )
                if response.status_code != 200:
                    raise HTTPException(status_code=401, detail="Invalid Google token")
                
                google_data = response.json()
                google_email = google_data.get("email", request.email).lower()
        except Exception as e:
            logger.error(f"Google token verification failed: {e}")
            google_data = {"email": google_email, "name": request.name}
    else:
        raise HTTPException(status_code=400, detail="Either google_id or id_token is required")
    
    # Create device_id for user
    user_device_id = f"dev_{uuid.uuid4().hex[:20]}"
    
    # Check if user exists
    user = await db.users.find_one({"email": google_email})
    
    if user:
        # Update auth provider, device_id and last login
        user_device_id = user.get("device_id", user_device_id)
        await db.users.update_one(
            {"_id": user["_id"]},
            {"$set": {
                "auth_provider": "google",
                "name": request.name or user.get("name", ""),
                "picture": request.picture or user.get("picture"),
                "device_id": user_device_id,
                "last_login": datetime.now(timezone.utc).isoformat()
            }}
        )
    else:
        # Create new user
        user_id = str(uuid.uuid4())
        user = {
            "_id": user_id,
            "email": google_email,
            "name": request.name,
            "picture": request.picture,
            "password_hash": None,
            "phone": None,
            "auth_provider": "google",
            "account_type": "free",
            "subscription_expires_at": None,
            "is_email_verified": True,  # Google emails are verified
            "device_id": user_device_id,
            "created_at": datetime.now(timezone.utc).isoformat(),
            "last_login": datetime.now(timezone.utc).isoformat()
        }
        await db.users.insert_one(user)
    
    # Generate tokens
    access_token = create_access_token(user["_id"], user["email"])
    refresh_token = create_refresh_token(user["_id"])
    
    return {
        "access_token": access_token,
        "refresh_token": refresh_token,
        "token_type": "bearer",
        "user": {
            "id": user["_id"],
            "email": user["email"],
            "name": user.get("name", ""),
            "phone": user.get("phone"),
            "auth_provider": "google",
            "account_type": user.get("account_type", "free"),
            "is_email_verified": True
        }
    }


@api_router.post("/auth/apple")
async def apple_auth(request: AppleAuthRequest):
    """Authenticate with Apple."""
    apple_email = request.email.lower()
    
    # Create device_id for user
    user_device_id = f"dev_{uuid.uuid4().hex[:20]}"
    
    # Check if user exists
    user = await db.users.find_one({"email": apple_email})
    
    if user:
        user_device_id = user.get("device_id", user_device_id)
        await db.users.update_one(
            {"_id": user["_id"]},
            {"$set": {
                "auth_provider": "apple",
                "name": request.name or user.get("name", ""),
                "device_id": user_device_id,
                "last_login": datetime.now(timezone.utc).isoformat()
            }}
        )
    else:
        user_id = str(uuid.uuid4())
        user = {
            "_id": user_id,
            "email": apple_email,
            "name": request.name,
            "password_hash": None,
            "phone": None,
            "auth_provider": "apple",
            "account_type": "free",
            "subscription_expires_at": None,
            "is_email_verified": True,
            "device_id": user_device_id,
            "created_at": datetime.now(timezone.utc).isoformat(),
            "last_login": datetime.now(timezone.utc).isoformat()
        }
        await db.users.insert_one(user)
    
    access_token = create_access_token(user["_id"], user["email"])
    refresh_token = create_refresh_token(user["_id"])
    
    return {
        "access_token": access_token,
        "refresh_token": refresh_token,
        "token_type": "bearer",
        "device_id": user_device_id,
        "user": {
            "id": user["_id"],
            "email": user["email"],
            "name": user.get("name", ""),
            "phone": user.get("phone"),
            "auth_provider": "apple",
            "account_type": user.get("account_type", "free"),
            "is_email_verified": True,
            "device_id": user_device_id
        }
    }


@api_router.post("/auth/facebook")
async def facebook_auth(request: FacebookAuthRequest):
    """Authenticate with Facebook."""
    fb_email = request.email.lower()
    fb_name = request.name
    
    # If facebook_id is provided directly (from mobile app), use it
    if request.facebook_id:
        # The user is authenticated via Facebook OAuth on mobile
        fb_data = {
            "id": request.facebook_id,
            "email": fb_email,
            "name": fb_name,
            "picture": request.picture
        }
    elif request.access_token:
        # Verify Facebook access token (for web flow)
        try:
            async with httpx.AsyncClient() as client:
                response = await client.get(
                    "https://graph.facebook.com/me",
                    params={
                        "access_token": request.access_token,
                        "fields": "id,email,name,picture"
                    }
                )
                if response.status_code != 200:
                    raise HTTPException(status_code=401, detail="Invalid Facebook token")
                
                fb_data = response.json()
                fb_email = fb_data.get("email", request.email).lower()
                fb_name = fb_data.get("name", request.name)
        except Exception as e:
            logger.error(f"Facebook token verification failed: {e}")
            fb_data = {"email": fb_email, "name": fb_name}
    else:
        raise HTTPException(status_code=400, detail="Either facebook_id or access_token is required")
    
    # Create device_id for user
    user_device_id = f"dev_{uuid.uuid4().hex[:20]}"
    
    # Check if user exists
    user = await db.users.find_one({"email": fb_email})
    
    if user:
        user_device_id = user.get("device_id", user_device_id)
        await db.users.update_one(
            {"_id": user["_id"]},
            {"$set": {
                "auth_provider": "facebook",
                "name": fb_name or user.get("name", ""),
                "picture": request.picture or user.get("picture"),
                "device_id": user_device_id,
                "last_login": datetime.now(timezone.utc).isoformat()
            }}
        )
    else:
        user_id = str(uuid.uuid4())
        user = {
            "_id": user_id,
            "email": fb_email,
            "name": fb_name,
            "picture": request.picture,
            "password_hash": None,
            "phone": None,
            "auth_provider": "facebook",
            "account_type": "free",
            "subscription_expires_at": None,
            "is_email_verified": True,  # Facebook emails are verified
            "device_id": user_device_id,
            "created_at": datetime.now(timezone.utc).isoformat(),
            "last_login": datetime.now(timezone.utc).isoformat()
        }
        await db.users.insert_one(user)
    
    access_token = create_access_token(user["_id"], user["email"])
    refresh_token = create_refresh_token(user["_id"])
    
    return {
        "access_token": access_token,
        "refresh_token": refresh_token,
        "token_type": "bearer",
        "device_id": user_device_id,
        "user": {
            "id": user["_id"],
            "email": user["email"],
            "name": user.get("name", ""),
            "phone": user.get("phone"),
            "auth_provider": "facebook",
            "account_type": user.get("account_type", "free"),
            "is_email_verified": True,
            "device_id": user_device_id
        }
    }


@api_router.post("/auth/phone/request-otp")
async def request_phone_otp(request: PhoneOTPRequest):
    """Request OTP for phone authentication."""
    import random
    
    # Generate 6-digit OTP
    otp = "".join([str(random.randint(0, 9)) for _ in range(6)])
    
    # Store OTP in database (expires in 10 minutes)
    await db.phone_otps.delete_many({"phone_number": request.phone_number})  # Remove old OTPs
    await db.phone_otps.insert_one({
        "_id": str(uuid.uuid4()),
        "phone_number": request.phone_number,
        "otp": otp,
        "created_at": datetime.now(timezone.utc).isoformat(),
        "expires_at": (datetime.now(timezone.utc) + timedelta(minutes=10)).isoformat()
    })
    
    # TODO: Send SMS via Twilio in production
    # For now, log the OTP (MOCK)
    logger.info(f"📱 [MOCK SMS] OTP for {request.phone_number}: {otp}")
    
    return {
        "success": True,
        "message": "OTP sent successfully",
        "mock_otp": otp  # Remove in production!
    }


@api_router.post("/auth/phone/verify-otp")
async def verify_phone_otp(request: PhoneOTPVerifyRequest):
    """Verify phone OTP."""
    # Find valid OTP
    otp_record = await db.phone_otps.find_one({
        "phone_number": request.phone_number,
        "otp": request.otp
    })
    
    if not otp_record:
        raise HTTPException(status_code=400, detail="Invalid OTP")
    
    # Check expiration
    expires_at = datetime.fromisoformat(otp_record["expires_at"].replace("Z", "+00:00"))
    if datetime.now(timezone.utc) > expires_at:
        raise HTTPException(status_code=400, detail="OTP expired")
    
    # Mark as verified (store temporarily)
    await db.phone_otps.delete_one({"_id": otp_record["_id"]})
    await db.pending_phone_auth.delete_many({"phone_number": request.phone_number})
    await db.pending_phone_auth.insert_one({
        "_id": str(uuid.uuid4()),
        "phone_number": request.phone_number,
        "verified_at": datetime.now(timezone.utc).isoformat(),
        "expires_at": (datetime.now(timezone.utc) + timedelta(minutes=15)).isoformat()
    })
    
    return {"success": True, "message": "OTP verified. Please provide your email."}


@api_router.post("/auth/phone/verify-firebase")
async def verify_firebase_phone(request: FirebasePhoneVerifyRequest):
    """Verify phone authentication via Firebase.
    
    This endpoint is called after Firebase phone auth succeeds.
    It marks the phone as verified in our system.
    """
    logger.info(f"🔥 Firebase phone verify: {request.phone_number} - UID: {request.firebase_uid}")
    
    # Store as verified (same as mock OTP flow)
    await db.pending_phone_auth.delete_many({"phone_number": request.phone_number})
    await db.pending_phone_auth.insert_one({
        "_id": str(uuid.uuid4()),
        "phone_number": request.phone_number,
        "firebase_uid": request.firebase_uid,
        "verified_at": datetime.now(timezone.utc).isoformat(),
        "expires_at": (datetime.now(timezone.utc) + timedelta(minutes=15)).isoformat()
    })
    
    return {"success": True, "message": "Phone verified via Firebase. Please provide your email."}


@api_router.post("/auth/phone/complete")
async def complete_phone_auth(request: PhoneCompleteRequest):
    """Complete phone authentication by providing email."""
    # Check pending verification
    pending = await db.pending_phone_auth.find_one({
        "phone_number": request.phone_number
    })
    
    if not pending:
        raise HTTPException(status_code=400, detail="Phone not verified. Please request OTP first.")
    
    expires_at = datetime.fromisoformat(pending["expires_at"].replace("Z", "+00:00"))
    if datetime.now(timezone.utc) > expires_at:
        raise HTTPException(status_code=400, detail="Verification expired. Please try again.")
    
    # Check if user with this phone exists
    user = await db.users.find_one({"phone": request.phone_number})
    
    if user:
        # Update email if needed
        await db.users.update_one(
            {"_id": user["_id"]},
            {"$set": {
                "email": request.email.lower(),
                "is_email_verified": True,
                "last_login": datetime.now(timezone.utc).isoformat()
            }}
        )
    else:
        # Check if email already exists
        existing_email = await db.users.find_one({"email": request.email.lower()})
        if existing_email:
            # Link phone to existing account
            await db.users.update_one(
                {"_id": existing_email["_id"]},
                {"$set": {
                    "phone": request.phone_number,
                    "last_login": datetime.now(timezone.utc).isoformat()
                }}
            )
            user = existing_email
            user["phone"] = request.phone_number
        else:
            # Create new user
            user_id = str(uuid.uuid4())
            user = {
                "_id": user_id,
                "email": request.email.lower(),
                "name": "",
                "password_hash": None,
                "phone": request.phone_number,
                "auth_provider": "phone",
                "account_type": "free",
                "subscription_expires_at": None,
                "is_email_verified": True,
                "created_at": datetime.now(timezone.utc).isoformat(),
                "last_login": datetime.now(timezone.utc).isoformat()
            }
            await db.users.insert_one(user)
    
    # Cleanup
    await db.pending_phone_auth.delete_one({"_id": pending["_id"]})
    
    access_token = create_access_token(user["_id"], user["email"])
    refresh_token = create_refresh_token(user["_id"])
    
    return {
        "access_token": access_token,
        "refresh_token": refresh_token,
        "token_type": "bearer",
        "user": {
            "id": user["_id"],
            "email": user["email"],
            "name": user.get("name", ""),
            "phone": user.get("phone"),
            "auth_provider": "phone",
            "account_type": user.get("account_type", "free"),
            "is_email_verified": True
        }
    }


@api_router.post("/auth/refresh")
async def refresh_token(refresh_token: str = Body(..., embed=True)):
    """Refresh access token."""
    try:
        payload = verify_token(refresh_token)
        if payload.get("type") != "refresh":
            raise HTTPException(status_code=401, detail="Invalid refresh token")
        
        user_id = payload.get("sub")
        user = await db.users.find_one({"_id": user_id})
        
        if not user:
            raise HTTPException(status_code=401, detail="User not found")
        
        new_access_token = create_access_token(user["_id"], user["email"])
        new_refresh_token = create_refresh_token(user["_id"])
        
        return {
            "access_token": new_access_token,
            "refresh_token": new_refresh_token,
            "token_type": "bearer"
        }
    except JWTError:
        raise HTTPException(status_code=401, detail="Invalid refresh token")


@api_router.get("/auth/me")
async def get_me(user: dict = Depends(get_current_user)):
    """Get current user profile."""
    # Check subscription status
    account_type = user.get("account_type", "free")
    subscription_expires = user.get("subscription_expires_at")
    
    if subscription_expires:
        expires_at = datetime.fromisoformat(subscription_expires.replace("Z", "+00:00"))
        if datetime.now(timezone.utc) > expires_at:
            account_type = "free"
            await db.users.update_one(
                {"_id": user["_id"]},
                {"$set": {"account_type": "free", "subscription_expires_at": None}}
            )
    
    return {
        "id": user["_id"],
        "email": user["email"],
        "name": user.get("name", ""),
        "phone": user.get("phone"),
        "auth_provider": user.get("auth_provider", "email"),
        "account_type": account_type,
        "subscription_expires_at": subscription_expires,
        "is_email_verified": user.get("is_email_verified", False),
        "created_at": user.get("created_at")
    }


@api_router.post("/auth/update-phone")
async def update_phone(request: UpdatePhoneRequest, user: dict = Depends(get_current_user)):
    """Add or update phone number for current user."""
    await db.users.update_one(
        {"_id": user["_id"]},
        {"$set": {"phone": request.phone_number}}
    )
    return {"success": True, "message": "Phone number updated"}


@api_router.post("/auth/logout")
async def logout(user: dict = Depends(get_current_user)):
    """Logout current user."""
    # In production, you might want to blacklist the token
    return {"success": True, "message": "Logged out successfully"}


# ============ PROMO CODES ============

@api_router.post("/auth/apply-promo")
async def apply_promo_code(request: PromoCodeRequest, user: dict = Depends(get_current_user)):
    """Apply a promo code for free premium access."""
    code_upper = request.code.upper()
    
    # Find promo code
    promo = await db.promo_codes.find_one({
        "code": code_upper,
        "is_active": True
    })
    
    if not promo:
        raise HTTPException(status_code=400, detail="Invalid promo code")
    
    # Check if already used max times
    if promo.get("max_uses") and promo.get("used_count", 0) >= promo["max_uses"]:
        raise HTTPException(status_code=400, detail="Promo code has expired")
    
    # Check if user already used this code
    user_uses = await db.promo_code_uses.find_one({
        "user_id": user["_id"],
        "code": code_upper
    })
    if user_uses:
        raise HTTPException(status_code=400, detail="You have already used this promo code")
    
    # Calculate new subscription end date
    duration_days = promo.get("duration_days", 30)
    current_expires = user.get("subscription_expires_at")
    
    if current_expires:
        # Extend existing subscription
        base_date = datetime.fromisoformat(current_expires.replace("Z", "+00:00"))
        if base_date < datetime.now(timezone.utc):
            base_date = datetime.now(timezone.utc)
    else:
        base_date = datetime.now(timezone.utc)
    
    new_expires = base_date + timedelta(days=duration_days)
    
    # Update user
    await db.users.update_one(
        {"_id": user["_id"]},
        {"$set": {
            "account_type": "paid",
            "subscription_expires_at": new_expires.isoformat()
        }}
    )
    
    # Record promo use
    await db.promo_code_uses.insert_one({
        "_id": str(uuid.uuid4()),
        "user_id": user["_id"],
        "code": code_upper,
        "used_at": datetime.now(timezone.utc).isoformat()
    })
    
    # Increment used count
    await db.promo_codes.update_one(
        {"code": code_upper},
        {"$inc": {"used_count": 1}}
    )
    
    return {
        "success": True,
        "message": f"Promo code applied! You have {duration_days} days of premium access.",
        "account_type": "paid",
        "subscription_expires_at": new_expires.isoformat()
    }


# ============ ADMIN - PROMO CODE MANAGEMENT ============

@api_router.post("/admin/promo-codes")
async def create_promo_code(
    code: str = Body(...),
    duration_days: int = Body(30),
    max_uses: Optional[int] = Body(None),
    admin_key: str = Query(None),
    admin_token: str = Query(None)
):
    """Create a new promo code (Admin only)."""
    is_valid = admin_key == ADMIN_PASSWORD
    if not is_valid and admin_token:
        is_valid = await verify_admin_token(admin_token)
    if not is_valid:
        raise HTTPException(status_code=403, detail="Unauthorized")
    
    code_upper = code.upper()
    
    # Check if code exists
    existing = await db.promo_codes.find_one({"code": code_upper})
    if existing:
        raise HTTPException(status_code=400, detail="Promo code already exists")
    
    promo = {
        "_id": str(uuid.uuid4()),
        "code": code_upper,
        "duration_days": duration_days,
        "max_uses": max_uses,
        "used_count": 0,
        "is_active": True,
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    
    await db.promo_codes.insert_one(promo)
    
    return {"success": True, "promo_code": promo}


@api_router.get("/admin/promo-codes")
async def list_promo_codes(
    admin_key: str = Query(None),
    admin_token: str = Query(None)
):
    """List all promo codes (Admin only)."""
    is_valid = admin_key == ADMIN_PASSWORD
    if not is_valid and admin_token:
        is_valid = await verify_admin_token(admin_token)
    if not is_valid:
        raise HTTPException(status_code=403, detail="Unauthorized")
    
    codes = await db.promo_codes.find().to_list(100)
    return {"promo_codes": codes}


@api_router.delete("/admin/promo-codes/{code}")
async def delete_promo_code(
    code: str, 
    admin_key: str = Query(None),
    admin_token: str = Query(None)
):
    """Delete a promo code (Admin only)."""
    is_valid = admin_key == ADMIN_PASSWORD
    if not is_valid and admin_token:
        is_valid = await verify_admin_token(admin_token)
    if not is_valid:
        raise HTTPException(status_code=403, detail="Unauthorized")
    
    result = await db.promo_codes.delete_one({"code": code.upper()})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Promo code not found")
    
    return {"success": True, "message": "Promo code deleted"}


@api_router.patch("/admin/promo-codes/{code}/toggle")
async def toggle_promo_code(
    code: str, 
    is_active: bool = Query(...), 
    admin_key: str = Query(None),
    admin_token: str = Query(None)
):
    """Toggle a promo code active/inactive (Admin only)."""
    is_valid = admin_key == ADMIN_PASSWORD
    if not is_valid and admin_token:
        is_valid = await verify_admin_token(admin_token)
    if not is_valid:
        raise HTTPException(status_code=403, detail="Unauthorized")
    
    result = await db.promo_codes.update_one(
        {"code": code.upper()},
        {"$set": {"is_active": is_active}}
    )
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Promo code not found")
    
    return {"success": True, "is_active": is_active}


@api_router.get("/admin/users")
async def list_users(admin_key: str = Query(...), skip: int = 0, limit: int = 100):
    """List all registered users (Admin only)."""
    if admin_key != ADMIN_PASSWORD:
        raise HTTPException(status_code=403, detail="Unauthorized")
    
    users = await db.users.find().sort("created_at", -1).skip(skip).limit(limit).to_list(limit)
    return {"users": users}


@api_router.post("/admin/users/{user_id}/upgrade")
async def upgrade_user(user_id: str, days: int = Body(..., embed=True), admin_key: str = Query(...)):
    """Upgrade a user to premium (Admin only)."""
    if admin_key != ADMIN_PASSWORD:
        raise HTTPException(status_code=403, detail="Unauthorized")
    
    user = await db.users.find_one({"_id": user_id})
    if not user:
        raise HTTPException(status_code=404, detail="User not found")
    
    # Calculate new expiration
    current_expires = user.get("subscription_expires_at")
    if current_expires:
        base_date = datetime.fromisoformat(current_expires.replace("Z", "+00:00"))
        if base_date < datetime.now(timezone.utc):
            base_date = datetime.now(timezone.utc)
    else:
        base_date = datetime.now(timezone.utc)
    
    new_expires = base_date + timedelta(days=days)
    
    await db.users.update_one(
        {"_id": user_id},
        {"$set": {
            "account_type": "paid",
            "subscription_expires_at": new_expires.isoformat()
        }}
    )
    
    return {"success": True, "subscription_expires_at": new_expires.isoformat()}


@api_router.post("/admin/users/{user_id}/downgrade")
async def downgrade_user(user_id: str, admin_key: str = Query(...)):
    """Downgrade a user to free (Admin only)."""
    if admin_key != ADMIN_PASSWORD:
        raise HTTPException(status_code=403, detail="Unauthorized")
    
    result = await db.users.update_one(
        {"_id": user_id},
        {"$set": {
            "account_type": "free",
            "subscription_expires_at": None
        }}
    )
    
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="User not found")
    
    return {"success": True}


# ============ DATA EXPORT (PAID USERS ONLY) ============

def check_user_is_paid(user: dict) -> bool:
    """Check if user has active premium subscription."""
    if user.get("account_type") != "paid":
        return False
    
    # Check if subscription is still valid
    expires_at = user.get("subscription_expires_at")
    if expires_at:
        try:
            expires = datetime.fromisoformat(expires_at.replace("Z", "+00:00"))
            if datetime.now(timezone.utc) > expires:
                return False
        except:
            pass
    
    return True


@api_router.get("/export/receipts")
async def export_receipts_excel(user: dict = Depends(get_current_user)):
    """Export user's receipts as Excel file (Paid users only)."""
    
    # Check if user is paid
    if not check_user_is_paid(user):
        raise HTTPException(
            status_code=403, 
            detail="Η εξαγωγή δεδομένων είναι διαθέσιμη μόνο για συνδρομητές apodixxi+"
        )
    
    # Get user's receipts
    user_email = user.get("email")
    
    # Get all receipts (in production, filter by user's devices/ownership)
    receipts = await db.receipts.find().sort("date", -1).to_list(1000)
    
    if not receipts:
        raise HTTPException(status_code=404, detail="Δεν βρέθηκαν αποδείξεις")
    
    # Create Excel workbook
    wb = Workbook()
    
    # ===== Sheet 1: Receipts Summary =====
    ws1 = wb.active
    ws1.title = "Αποδείξεις"
    
    # Header styling
    header_fill = PatternFill(start_color="0D9488", end_color="0D9488", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Headers
    headers = ["Ημερομηνία", "Κατάστημα", "ΑΦΜ", "Αρ. Απόδειξης", "Σύνολο (€)", "Αριθμός Προϊόντων"]
    for col, header in enumerate(headers, 1):
        cell = ws1.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
        cell.border = border
    
    # Data rows
    for row, receipt in enumerate(receipts, 2):
        ws1.cell(row=row, column=1, value=receipt.get("date", "")).border = border
        ws1.cell(row=row, column=2, value=receipt.get("store_name", "")).border = border
        ws1.cell(row=row, column=3, value=receipt.get("store_vat", "")).border = border
        ws1.cell(row=row, column=4, value=receipt.get("receipt_number", "")).border = border
        total = receipt.get("total", 0)
        ws1.cell(row=row, column=5, value=float(total) if total else 0).border = border
        ws1.cell(row=row, column=5).number_format = '#,##0.00'
        items_count = len(receipt.get("items", []))
        ws1.cell(row=row, column=6, value=items_count).border = border
    
    # Adjust column widths
    ws1.column_dimensions['A'].width = 15
    ws1.column_dimensions['B'].width = 25
    ws1.column_dimensions['C'].width = 15
    ws1.column_dimensions['D'].width = 20
    ws1.column_dimensions['E'].width = 12
    ws1.column_dimensions['F'].width = 18
    
    # ===== Sheet 2: All Products =====
    ws2 = wb.create_sheet("Προϊόντα")
    
    # Headers
    product_headers = ["Ημερομηνία", "Κατάστημα", "Προϊόν", "Ποσότητα", "Τιμή Μονάδας (€)", "Σύνολο (€)"]
    for col, header in enumerate(product_headers, 1):
        cell = ws2.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
        cell.border = border
    
    # Data rows
    row_num = 2
    for receipt in receipts:
        date = receipt.get("date", "")
        store = receipt.get("store_name", "")
        for item in receipt.get("items", []):
            ws2.cell(row=row_num, column=1, value=date).border = border
            ws2.cell(row=row_num, column=2, value=store).border = border
            ws2.cell(row=row_num, column=3, value=item.get("description", "")).border = border
            qty = item.get("quantity", 1)
            ws2.cell(row=row_num, column=4, value=float(qty) if qty else 1).border = border
            unit_price = item.get("unit_price", 0)
            ws2.cell(row=row_num, column=5, value=float(unit_price) if unit_price else 0).border = border
            ws2.cell(row=row_num, column=5).number_format = '#,##0.00'
            total_price = item.get("total_price", 0)
            ws2.cell(row=row_num, column=6, value=float(total_price) if total_price else 0).border = border
            ws2.cell(row=row_num, column=6).number_format = '#,##0.00'
            row_num += 1
    
    # Adjust column widths
    ws2.column_dimensions['A'].width = 15
    ws2.column_dimensions['B'].width = 25
    ws2.column_dimensions['C'].width = 40
    ws2.column_dimensions['D'].width = 12
    ws2.column_dimensions['E'].width = 18
    ws2.column_dimensions['F'].width = 12
    
    # ===== Sheet 3: Statistics =====
    ws3 = wb.create_sheet("Στατιστικά")
    
    # Calculate stats
    total_amount = sum(float(r.get("total", 0) or 0) for r in receipts)
    total_items = sum(len(r.get("items", [])) for r in receipts)
    
    # Store breakdown
    store_totals = {}
    for r in receipts:
        store = r.get("store_name", "Άγνωστο")
        amount = float(r.get("total", 0) or 0)
        if store in store_totals:
            store_totals[store]["count"] += 1
            store_totals[store]["amount"] += amount
        else:
            store_totals[store] = {"count": 1, "amount": amount}
    
    # Headers
    ws3.cell(row=1, column=1, value="Συνολικά Στατιστικά").font = Font(bold=True, size=14)
    ws3.cell(row=3, column=1, value="Σύνολο Αποδείξεων:")
    ws3.cell(row=3, column=2, value=len(receipts))
    ws3.cell(row=4, column=1, value="Σύνολο Προϊόντων:")
    ws3.cell(row=4, column=2, value=total_items)
    ws3.cell(row=5, column=1, value="Συνολικά Έξοδα:")
    ws3.cell(row=5, column=2, value=total_amount).number_format = '#,##0.00 €'
    
    ws3.cell(row=7, column=1, value="Ανάλυση ανά Κατάστημα").font = Font(bold=True, size=12)
    
    stat_headers = ["Κατάστημα", "Αριθμός Αποδείξεων", "Σύνολο (€)"]
    for col, header in enumerate(stat_headers, 1):
        cell = ws3.cell(row=8, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
    
    row_num = 9
    for store, data in sorted(store_totals.items(), key=lambda x: x[1]["amount"], reverse=True):
        ws3.cell(row=row_num, column=1, value=store).border = border
        ws3.cell(row=row_num, column=2, value=data["count"]).border = border
        ws3.cell(row=row_num, column=3, value=data["amount"]).border = border
        ws3.cell(row=row_num, column=3).number_format = '#,##0.00'
        row_num += 1
    
    ws3.column_dimensions['A'].width = 30
    ws3.column_dimensions['B'].width = 20
    ws3.column_dimensions['C'].width = 15
    
    # Save to bytes buffer
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    # Generate filename with date
    filename = f"apodixxi_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


@api_router.get("/export/check-access")
async def check_export_access(user: dict = Depends(get_current_user)):
    """Check if user can access export feature."""
    is_paid = check_user_is_paid(user)
    
    return {
        "can_export": is_paid,
        "account_type": user.get("account_type", "free"),
        "subscription_expires_at": user.get("subscription_expires_at"),
        "message": None if is_paid else "Η εξαγωγή δεδομένων είναι διαθέσιμη μόνο για συνδρομητές apodixxi+"
    }



# ============ RECOMMENDATIONS SYSTEM ============

class PromotionCreate(BaseModel):
    title: str
    description: str = ""
    product_name: str = ""
    price: Optional[float] = None
    original_price: Optional[float] = None
    store_name: str = ""
    store_vat: str = ""
    image_url: str = ""
    barcode_code: str = ""
    barcode_image_url: str = ""
    target_categories: List[str] = []
    target_stores: List[str] = []
    target_all_users: bool = True
    start_date: Optional[str] = None
    end_date: Optional[str] = None
    priority: int = 0  # Higher = shown first


@api_router.get("/recommendations")
async def get_recommendations(
    device_id: str = Query(...),
    limit: int = Query(5, ge=1, le=20),
    location: str = Query("dashboard")  # dashboard, after_save, compare
):
    """Get personalized recommendations for a user."""
    recommendations = []
    
    # 1. Get active admin promotions
    now = datetime.now(timezone.utc).isoformat()
    promotions = await db.promotions.find({
        "is_active": True,
        "$or": [
            {"start_date": None},
            {"start_date": {"$lte": now}}
        ]
    }).sort("priority", -1).limit(limit).to_list(limit)
    
    for promo in promotions:
        # Check end date
        if promo.get("end_date") and promo["end_date"] < now:
            continue
        
        recommendations.append({
            "id": promo["_id"],
            "type": "promotion",
            "title": promo.get("title", ""),
            "description": promo.get("description", ""),
            "product_name": promo.get("product_name", ""),
            "price": promo.get("price"),
            "original_price": promo.get("original_price"),
            "store_name": promo.get("store_name", ""),
            "image_url": promo.get("image_url", ""),
            "barcode_code": promo.get("barcode_code", ""),
            "barcode_image_url": promo.get("barcode_image_url", ""),
            "is_sponsored": True
        })
    
    # 2. Generate automatic recommendations based on purchase history
    if len(recommendations) < limit:
        # Get user's frequent products
        pipeline = [
            {"$match": {"device_id": device_id}},
            {"$unwind": "$items"},
            {"$group": {
                "_id": {"$toUpper": "$items.description"},
                "count": {"$sum": 1},
                "avg_price": {"$avg": "$items.unit_price"},
                "last_store": {"$last": "$store_name"},
                "last_price": {"$last": "$items.unit_price"}
            }},
            {"$match": {"count": {"$gte": 2}}},  # Products bought at least twice
            {"$sort": {"count": -1}},
            {"$limit": 10}
        ]
        frequent_products = await db.receipts.aggregate(pipeline).to_list(10)
        
        # Find better prices for frequent products
        for product in frequent_products:
            if len(recommendations) >= limit:
                break
            
            product_name = product["_id"]
            last_price = product.get("last_price", 0)
            last_store = product.get("last_store", "")
            
            # Search for this product in other receipts with lower price
            better_price = await db.receipts.find_one({
                "items.description": {"$regex": product_name[:20], "$options": "i"},
                "store_name": {"$ne": last_store}
            }, sort=[("items.unit_price", 1)])
            
            if better_price:
                for item in better_price.get("items", []):
                    if product_name[:15].upper() in item.get("description", "").upper():
                        item_price = item.get("unit_price", 0)
                        if item_price and item_price < last_price * 0.95:  # At least 5% cheaper
                            savings = round(last_price - item_price, 2)
                            recommendations.append({
                                "id": f"auto_{product_name[:10]}",
                                "type": "price_alert",
                                "title": f"Εξοικονόμηση €{savings}!",
                                "description": f"Το {product_name[:30]} είναι φθηνότερο στο {better_price.get('store_name', '')}",
                                "product_name": product_name,
                                "price": item_price,
                                "original_price": last_price,
                                "store_name": better_price.get("store_name", ""),
                                "is_sponsored": False
                            })
                            break
    
    # 3. Add store-based recommendations
    if len(recommendations) < limit and location == "dashboard":
        # Get user's most visited store
        store_pipeline = [
            {"$match": {"device_id": device_id}},
            {"$group": {"_id": "$store_name", "count": {"$sum": 1}}},
            {"$sort": {"count": -1}},
            {"$limit": 1}
        ]
        top_store = await db.receipts.aggregate(store_pipeline).to_list(1)
        
        if top_store:
            store_name = top_store[0]["_id"]
            visit_count = top_store[0]["count"]
            
            recommendations.append({
                "id": f"store_{store_name[:10]}",
                "type": "insight",
                "title": f"Τακτικός πελάτης {store_name}",
                "description": f"Έχετε {visit_count} επισκέψεις. Δείτε τις τελευταίες προσφορές!",
                "store_name": store_name,
                "is_sponsored": False
            })
    
    return {
        "recommendations": recommendations[:limit],
        "total": len(recommendations)
    }


@api_router.get("/recommendations/after-save")
async def get_after_save_recommendations(
    device_id: str = Query(...),
    receipt_id: str = Query(...),
    limit: int = Query(3)
):
    """Get recommendations after saving a receipt."""
    recommendations = []
    
    # Get the saved receipt
    receipt = await db.receipts.find_one({"id": receipt_id})
    if not receipt:
        return {"recommendations": [], "total": 0}
    
    store_name = receipt.get("store_name", "")
    items = receipt.get("items", [])
    
    # Find items that could be cheaper elsewhere
    for item in items[:5]:  # Check first 5 items
        item_desc = item.get("description", "")
        item_price = item.get("unit_price", 0)
        
        if not item_desc or not item_price:
            continue
        
        # Search for cheaper alternatives
        search_term = item_desc[:20].upper()
        
        pipeline = [
            {"$unwind": "$items"},
            {"$match": {
                "store_name": {"$ne": store_name},
                "items.description": {"$regex": search_term, "$options": "i"},
                "items.unit_price": {"$lt": item_price * 0.9}  # At least 10% cheaper
            }},
            {"$sort": {"items.unit_price": 1}},
            {"$limit": 1}
        ]
        
        cheaper = await db.receipts.aggregate(pipeline).to_list(1)
        
        if cheaper and len(recommendations) < limit:
            cheaper_item = cheaper[0]
            cheaper_price = cheaper_item["items"]["unit_price"]
            savings = round(item_price - cheaper_price, 2)
            
            recommendations.append({
                "id": f"save_{item_desc[:10]}",
                "type": "price_comparison",
                "title": f"Ήξερες ότι...",
                "description": f"Το \"{item_desc[:25]}\" είναι €{savings} φθηνότερο στο {cheaper_item.get('store_name', '')}",
                "product_name": item_desc,
                "price": cheaper_price,
                "original_price": item_price,
                "store_name": cheaper_item.get("store_name", ""),
                "savings": savings,
                "is_sponsored": False
            })
    
    return {
        "recommendations": recommendations,
        "total": len(recommendations)
    }


# ============ ADMIN PROMOTIONS MANAGEMENT ============

@api_router.post("/admin/promotions")
async def create_promotion(
    promo: PromotionCreate,
    admin_key: str = Query(None),
    admin_token: str = Query(None)
):
    """Create a new promotion (Admin only)."""
    is_valid = admin_key == ADMIN_PASSWORD
    if not is_valid and admin_token:
        is_valid = await verify_admin_token(admin_token)
    if not is_valid:
        raise HTTPException(status_code=403, detail="Unauthorized")
    
    promotion = {
        "_id": str(uuid.uuid4()),
        "title": promo.title,
        "description": promo.description,
        "product_name": promo.product_name,
        "price": promo.price,
        "original_price": promo.original_price,
        "store_name": promo.store_name,
        "store_vat": promo.store_vat,
        "image_url": promo.image_url,
        "barcode_code": promo.barcode_code,
        "barcode_image_url": promo.barcode_image_url,
        "target_categories": promo.target_categories,
        "target_stores": promo.target_stores,
        "target_all_users": promo.target_all_users,
        "start_date": promo.start_date,
        "end_date": promo.end_date,
        "priority": promo.priority,
        "is_active": True,
        "views_count": 0,
        "clicks_count": 0,
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    
    await db.promotions.insert_one(promotion)
    
    return {"success": True, "promotion": promotion}


@api_router.get("/admin/promotions")
async def list_promotions(
    admin_key: str = Query(None),
    admin_token: str = Query(None)
):
    """List all promotions (Admin only)."""
    is_valid = admin_key == ADMIN_PASSWORD
    if not is_valid and admin_token:
        is_valid = await verify_admin_token(admin_token)
    if not is_valid:
        raise HTTPException(status_code=403, detail="Unauthorized")
    
    promotions = await db.promotions.find().sort("created_at", -1).to_list(100)
    return {"promotions": promotions}


@api_router.patch("/admin/promotions/{promo_id}")
async def update_promotion(
    promo_id: str,
    updates: dict = Body(...),
    admin_key: str = Query(None),
    admin_token: str = Query(None)
):
    """Update a promotion (Admin only)."""
    is_valid = admin_key == ADMIN_PASSWORD
    if not is_valid and admin_token:
        is_valid = await verify_admin_token(admin_token)
    if not is_valid:
        raise HTTPException(status_code=403, detail="Unauthorized")
    
    # Remove _id from updates if present
    updates.pop("_id", None)
    updates["updated_at"] = datetime.now(timezone.utc).isoformat()
    
    result = await db.promotions.update_one(
        {"_id": promo_id},
        {"$set": updates}
    )
    
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Promotion not found")
    
    return {"success": True}


@api_router.delete("/admin/promotions/{promo_id}")
async def delete_promotion(
    promo_id: str, 
    admin_key: str = Query(None),
    admin_token: str = Query(None)
):
    """Delete a promotion (Admin only)."""
    is_valid = admin_key == ADMIN_PASSWORD
    if not is_valid and admin_token:
        is_valid = await verify_admin_token(admin_token)
    if not is_valid:
        raise HTTPException(status_code=403, detail="Unauthorized")
    
    result = await db.promotions.delete_one({"_id": promo_id})
    
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Promotion not found")
    
    return {"success": True}


@api_router.post("/recommendations/{rec_id}/click")
async def track_recommendation_click(rec_id: str):
    """Track when a user clicks on a recommendation."""
    await db.promotions.update_one(
        {"_id": rec_id},
        {"$inc": {"clicks_count": 1}}
    )
    return {"success": True}


@api_router.post("/recommendations/{rec_id}/view")
async def track_recommendation_view(rec_id: str):
    """Track when a recommendation is viewed."""
    await db.promotions.update_one(
        {"_id": rec_id},
        {"$inc": {"views_count": 1}}
    )
    return {"success": True}




@api_router.post("/devices/register")
async def register_device(input: DeviceRegister):
    existing = await db.devices.find_one({"device_id": input.device_id}, {"_id": 0})
    if existing:
        await db.devices.update_one({"device_id": input.device_id}, {"$set": {"language": input.language, "last_seen": datetime.now(timezone.utc).isoformat()}})
        return {"status": "updated", "device_id": input.device_id}
    doc = {"device_id": input.device_id, "language": input.language, "created_at": datetime.now(timezone.utc).isoformat(), "last_seen": datetime.now(timezone.utc).isoformat()}
    await db.devices.insert_one(doc)
    return {"status": "registered", "device_id": input.device_id}


@api_router.post("/receipts/import-url")
async def import_receipt_from_url(input: URLImportInput):
    url = input.url.strip()
    provider = detect_provider(url)

    if provider == 'epsilon_digital':
        # Check for duplicates even for epsilon digital
        if not input.force_import:
            existing = await db.receipts.find_one({"source_url": url, "device_id": input.device_id})
            if existing:
                existing.pop("_id", None)
                return {
                    "status": "duplicate",
                    "message": "Αυτή η απόδειξη έχει ήδη εισαχθεί.",
                    "existing_receipt": existing,
                    "url": url
                }
        return {"status": "webview_required", "url": url, "message": "This receipt requires WebView import. Opening in-app browser..."}

    if provider == 'unknown':
        raise HTTPException(status_code=400, detail="Unknown receipt provider. Supported: e-invoicing.gr, einvoice.impact.gr")

    # Check for duplicate receipt by URL
    if not input.force_import:
        existing = await db.receipts.find_one({"source_url": url, "device_id": input.device_id})
        if existing:
            existing.pop("_id", None)
            return {
                "status": "duplicate",
                "message": "Αυτή η απόδειξη έχει ήδη εισαχθεί.",
                "existing_receipt": existing,
                "url": url
            }

    if provider == 'entersoft':
        html = await fetch_entersoft_invoice(url)
        receipt_data = parse_entersoft(html, url)
    elif provider == 'impact':
        html = await fetch_html(url)
        receipt_data = parse_impact(html, url)
    else:
        raise HTTPException(status_code=400, detail="Parser not available for this provider")

    if not receipt_data["items"]:
        raise HTTPException(status_code=400, detail="Could not parse any products from this receipt")

    receipt = ReceiptData(device_id=input.device_id, **receipt_data)
    doc = receipt.dict()
    await db.receipts.insert_one(doc)

    # Index products
    for item in receipt_data["items"]:
        await db.products.update_one(
            {"description": item["description"], "store_name": receipt_data["store_name"]},
            {"$set": {
                "description": item["description"],
                "code": item["code"],
                "store_name": receipt_data["store_name"],
                "store_vat": receipt_data["store_vat"],
                "last_price": item["total_value"],
                "last_unit_price": item["unit_price"],
                "last_date": receipt_data["date"],
                "unit": item["unit"],
                "vat_percent": item["vat_percent"],
                "updated_at": datetime.now(timezone.utc).isoformat()
            },
            "$push": {"price_history": {"price": item["total_value"], "unit_price": item["unit_price"], "date": receipt_data["date"], "quantity": item["quantity"], "receipt_id": doc["id"]}}},
            upsert=True
        )

    doc.pop("_id", None)
    return {"status": "success", "receipt": doc}


@api_router.post("/receipts/import-xml")
async def import_receipt_from_xml(device_id: str = Form(...), file: UploadFile = File(...)):
    content = await file.read()
    xml_str = content.decode('utf-8')
    receipt_data = parse_mydata_xml(xml_str)

    if not receipt_data["items"]:
        raise HTTPException(status_code=400, detail="Could not parse any products from this XML")

    receipt = ReceiptData(device_id=device_id, **receipt_data)
    doc = receipt.dict()
    await db.receipts.insert_one(doc)

    for item in receipt_data["items"]:
        await db.products.update_one(
            {"description": item["description"], "store_name": receipt_data["store_name"]},
            {"$set": {
                "description": item["description"],
                "code": item["code"],
                "store_name": receipt_data["store_name"],
                "store_vat": receipt_data["store_vat"],
                "last_price": item["total_value"],
                "last_unit_price": item["unit_price"],
                "last_date": receipt_data["date"],
                "unit": item["unit"],
                "vat_percent": item["vat_percent"],
                "updated_at": datetime.now(timezone.utc).isoformat()
            },
            "$push": {"price_history": {"price": item["total_value"], "unit_price": item["unit_price"], "date": receipt_data["date"], "quantity": item["quantity"], "receipt_id": doc["id"]}}},
            upsert=True
        )

    doc.pop("_id", None)
    return {"status": "success", "receipt": doc}


@api_router.post("/receipts/import-webview")
async def import_receipt_from_webview(input: WebViewExtractedData):
    """Import receipt from WebView DOM extraction (Epsilon Digital stores)."""
    receipt_data = parse_webview_extracted(
        raw_text=input.raw_text,
        items_from_dom=input.items,
        store_hint=input.store_name,
        source_url=input.url,
        found_final_total=input.found_final_total
    )

    if not receipt_data["items"]:
        raise HTTPException(status_code=400, detail="Could not parse any products from the extracted data")

    receipt = ReceiptData(device_id=input.device_id, **receipt_data)
    doc = receipt.dict()
    await db.receipts.insert_one(doc)

    for item in receipt_data["items"]:
        await db.products.update_one(
            {"description": item["description"], "store_name": receipt_data["store_name"]},
            {"$set": {
                "description": item["description"],
                "code": item["code"],
                "store_name": receipt_data["store_name"],
                "store_vat": receipt_data["store_vat"],
                "last_price": item["total_value"],
                "last_unit_price": item["unit_price"],
                "last_date": receipt_data["date"],
                "unit": item["unit"],
                "updated_at": datetime.now(timezone.utc).isoformat()
            },
            "$push": {"price_history": {"price": item["total_value"], "unit_price": item["unit_price"], "date": receipt_data["date"], "quantity": item["quantity"], "receipt_id": doc["id"]}}},
            upsert=True
        )

    doc.pop("_id", None)
    return {"status": "success", "receipt": doc}


@api_router.post("/receipts/manual")
async def create_manual_receipt(input: ManualReceiptInput):
    receipt_data = {
        "store_name": input.store_name,
        "store_address": "",
        "store_vat": "",
        "receipt_number": "",
        "date": input.date,
        "payment_method": input.payment_method,
        "items": [item.dict() for item in input.items],
        "subtotal": input.total,
        "discount_total": 0.0,
        "total": input.total,
        "vat_total": 0.0,
        "net_total": input.total,
        "source_url": "",
        "source_type": "manual",
        "provider": "Manual Entry"
    }
    receipt = ReceiptData(device_id=input.device_id, **receipt_data)
    doc = receipt.dict()
    await db.receipts.insert_one(doc)

    for item in input.items:
        await db.products.update_one(
            {"description": item.description, "store_name": input.store_name},
            {"$set": {
                "description": item.description,
                "code": item.code,
                "store_name": input.store_name,
                "last_price": item.total_value,
                "last_unit_price": item.unit_price,
                "last_date": input.date,
                "unit": item.unit,
                "updated_at": datetime.now(timezone.utc).isoformat()
            },
            "$push": {"price_history": {"price": item.total_value, "unit_price": item.unit_price, "date": input.date, "quantity": item.quantity, "receipt_id": doc["id"]}}},
            upsert=True
        )

    doc.pop("_id", None)
    return {"status": "success", "receipt": doc}


@api_router.get("/receipts")
async def get_receipts(device_id: str = Query(...), skip: int = 0, limit: int = 50, search: str = ""):
    query = {"device_id": device_id}
    if search:
        query["$or"] = [
            {"store_name": {"$regex": search, "$options": "i"}},
            {"items.description": {"$regex": search, "$options": "i"}}
        ]
    cursor = db.receipts.find(query, {"_id": 0}).sort("created_at", -1).skip(skip).limit(limit)
    receipts = await cursor.to_list(limit)
    # Sanitize all float values to prevent JSON serialization errors
    sanitized_receipts = [sanitize_receipt_data(r) for r in receipts]
    total = await db.receipts.count_documents(query)
    return {"receipts": sanitized_receipts, "total": total}


# IMPORTANT: This must come BEFORE /receipts/{receipt_id} to avoid path conflict
@api_router.get("/receipts/by-store")
async def get_receipts_by_store(
    device_id: str = Query(...),
    store_name: str = Query(...),
    skip: int = Query(0, ge=0),
    limit: int = Query(100, ge=1, le=500)
):
    """Get all receipts from a specific store."""
    query = {
        "device_id": device_id,
        "store_name": {"$regex": f"^{store_name}$", "$options": "i"}
    }
    
    total = await db.receipts.count_documents(query)
    receipts = await db.receipts.find(query, {"_id": 0}).sort("created_at", -1).skip(skip).limit(limit).to_list(limit)
    sanitized_receipts = [sanitize_receipt_data(r) for r in receipts]
    
    return {
        "store_name": store_name,
        "total": total,
        "receipts": sanitized_receipts
    }


@api_router.get("/receipts/{receipt_id}")
async def get_receipt(receipt_id: str):
    receipt = await db.receipts.find_one({"id": receipt_id}, {"_id": 0})
    if not receipt:
        raise HTTPException(status_code=404, detail="Receipt not found")
    # Sanitize float values
    return sanitize_receipt_data(receipt)


@api_router.delete("/receipts/{receipt_id}")
async def delete_receipt(receipt_id: str):
    result = await db.receipts.delete_one({"id": receipt_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Receipt not found")
    return {"status": "deleted"}


@api_router.get("/products/search")
async def search_products(q: str = Query(..., min_length=2), device_id: str = Query("")):
    query = {"description": {"$regex": q, "$options": "i"}}
    products = await db.products.find(query, {"_id": 0}).limit(50).to_list(50)
    return {"products": products}


@api_router.get("/products/compare")
async def compare_product_prices(q: str = Query(..., min_length=2)):
    products = await db.products.find(
        {"description": {"$regex": q, "$options": "i"}},
        {"_id": 0}
    ).to_list(100)

    stores = {}
    for p in products:
        store = p.get("store_name", "Unknown")
        if store not in stores:
            stores[store] = []
        stores[store].append({
            "description": p["description"],
            "last_price": p.get("last_price", 0),
            "last_unit_price": p.get("last_unit_price", 0),
            "last_date": p.get("last_date", ""),
            "price_history": p.get("price_history", [])
        })

    return {"query": q, "stores": stores, "total_products": len(products)}


@api_router.get("/stats")
async def get_stats(device_id: str = Query(...)):
    total_receipts = await db.receipts.count_documents({"device_id": device_id})
    total_products = await db.products.count_documents({})

    pipeline = [
        {"$match": {"device_id": device_id}},
        {"$group": {"_id": None, "total_spent": {"$sum": "$total"}, "avg_receipt": {"$avg": "$total"}}}
    ]
    result = await db.receipts.aggregate(pipeline).to_list(1)
    total_spent = result[0]["total_spent"] if result else 0
    avg_receipt = result[0]["avg_receipt"] if result else 0

    store_pipeline = [
        {"$match": {"device_id": device_id}},
        {"$group": {"_id": "$store_name", "count": {"$sum": 1}, "total": {"$sum": "$total"}}},
        {"$sort": {"count": -1}}
    ]
    stores = await db.receipts.aggregate(store_pipeline).to_list(20)

    recent = await db.receipts.find({"device_id": device_id}, {"_id": 0, "id": 1, "store_name": 1, "total": 1, "date": 1, "created_at": 1}).sort("created_at", -1).limit(5).to_list(5)

    return {
        "total_receipts": total_receipts,
        "total_products": total_products,
        "total_spent": round(total_spent, 2),
        "avg_receipt": round(avg_receipt, 2),
        "stores": [{"name": s["_id"], "count": s["count"], "total": round(s["total"], 2)} for s in stores],
        "recent_receipts": recent
    }


@api_router.get("/stats/analytics")
async def get_analytics(device_id: str = Query(...), months: int = Query(default=6, ge=1, le=12)):
    """Get spending analytics: monthly breakdown and store distribution."""
    
    # Get all receipts for this device
    receipts = await db.receipts.find(
        {"device_id": device_id}, 
        {"_id": 0, "date": 1, "total": 1, "store_name": 1, "items": 1, "created_at": 1}
    ).to_list(10000)
    
    if not receipts:
        return {
            "monthly_spending": [],
            "store_distribution": [],
            "top_products": [],
            "spending_trend": "neutral",
            "total_this_month": 0,
            "total_last_month": 0,
            "change_percent": 0
        }
    
    # Parse dates and organize by month
    from collections import defaultdict
    monthly_data = defaultdict(float)
    store_totals = defaultdict(float)
    current_month_store_totals = defaultdict(float)  # Store distribution for current month
    product_counts = defaultdict(lambda: {"count": 0, "total": 0.0, "description": ""})
    
    now = datetime.now(timezone.utc)
    current_month_key = now.strftime("%Y-%m")
    last_month = (now.replace(day=1) - __import__('datetime').timedelta(days=1))
    last_month_key = last_month.strftime("%Y-%m")
    
    for receipt in receipts:
        total = safe_float(receipt.get("total", 0))
        store = receipt.get("store_name", "Unknown")
        
        # Try to parse date from receipt
        date_str = receipt.get("date", "") or receipt.get("created_at", "")
        month_key = None
        
        if date_str:
            # Try different date formats
            for fmt in ["%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y", "%d.%m.%Y"]:
                try:
                    parsed = datetime.strptime(date_str[:10], fmt)
                    month_key = parsed.strftime("%Y-%m")
                    break
                except (ValueError, TypeError):
                    continue
            
            # Try ISO format for created_at
            if not month_key and "T" in date_str:
                try:
                    month_key = date_str[:7]  # YYYY-MM
                except:
                    pass
        
        if month_key:
            monthly_data[month_key] += total
            # Track current month store distribution
            if month_key == current_month_key:
                current_month_store_totals[store] += total
        
        store_totals[store] += total
        
        # Count products
        for item in receipt.get("items", []):
            desc = item.get("description", "")
            if desc:
                product_counts[desc]["count"] += 1
                product_counts[desc]["total"] += safe_float(item.get("total_value", 0))
                product_counts[desc]["description"] = desc
    
    # Generate last N months for chart
    month_labels_gr = {
        "01": "Ιαν", "02": "Φεβ", "03": "Μαρ", "04": "Απρ",
        "05": "Μάι", "06": "Ιουν", "07": "Ιουλ", "08": "Αυγ",
        "09": "Σεπ", "10": "Οκτ", "11": "Νοε", "12": "Δεκ"
    }
    
    monthly_spending = []
    for i in range(months - 1, -1, -1):
        target_date = now.replace(day=1)
        for _ in range(i):
            target_date = (target_date - __import__('datetime').timedelta(days=1)).replace(day=1)
        
        month_key = target_date.strftime("%Y-%m")
        month_num = target_date.strftime("%m")
        month_label = month_labels_gr.get(month_num, month_num)
        
        monthly_spending.append({
            "month": month_key,
            "label": month_label,
            "amount": round(monthly_data.get(month_key, 0), 2)
        })
    
    # Store distribution (top 5 + others)
    sorted_stores = sorted(store_totals.items(), key=lambda x: x[1], reverse=True)
    store_colors = ["#0D9488", "#6366F1", "#F59E0B", "#EF4444", "#10B981", "#8B5CF6", "#EC4899", "#3B82F6"]
    
    store_distribution = []
    total_all_stores = sum(store_totals.values())
    others_total = 0
    
    for i, (store, total) in enumerate(sorted_stores):
        if i < 5:
            percentage = (total / total_all_stores * 100) if total_all_stores > 0 else 0
            store_distribution.append({
                "name": store,
                "amount": round(total, 2),
                "percentage": round(percentage, 1),
                "color": store_colors[i % len(store_colors)]
            })
        else:
            others_total += total
    
    if others_total > 0:
        percentage = (others_total / total_all_stores * 100) if total_all_stores > 0 else 0
        store_distribution.append({
            "name": "Άλλα",
            "amount": round(others_total, 2),
            "percentage": round(percentage, 1),
            "color": "#94A3B8"
        })
    
    # Top products
    sorted_products = sorted(product_counts.values(), key=lambda x: x["count"], reverse=True)[:5]
    top_products = [
        {"description": p["description"], "count": p["count"], "total": round(p["total"], 2)}
        for p in sorted_products
    ]
    
    # Spending trend
    this_month_total = monthly_data.get(current_month_key, 0)
    last_month_total = monthly_data.get(last_month_key, 0)
    
    if last_month_total > 0:
        change_percent = ((this_month_total - last_month_total) / last_month_total) * 100
    else:
        change_percent = 0
    
    if change_percent > 5:
        trend = "up"
    elif change_percent < -5:
        trend = "down"
    else:
        trend = "neutral"
    
    # Current month store distribution
    sorted_current_month_stores = sorted(current_month_store_totals.items(), key=lambda x: x[1], reverse=True)
    current_month_store_distribution = []
    total_current_month_stores = sum(current_month_store_totals.values())
    current_month_others_total = 0
    
    for i, (store, total) in enumerate(sorted_current_month_stores):
        if i < 5:
            percentage = (total / total_current_month_stores * 100) if total_current_month_stores > 0 else 0
            current_month_store_distribution.append({
                "name": store,
                "amount": round(total, 2),
                "percentage": round(percentage, 1),
                "color": store_colors[i % len(store_colors)]
            })
        else:
            current_month_others_total += total
    
    if current_month_others_total > 0:
        percentage = (current_month_others_total / total_current_month_stores * 100) if total_current_month_stores > 0 else 0
        current_month_store_distribution.append({
            "name": "Άλλα",
            "amount": round(current_month_others_total, 2),
            "percentage": round(percentage, 1),
            "color": "#94A3B8"
        })
    
    return {
        "monthly_spending": monthly_spending,
        "store_distribution": store_distribution,
        "current_month_store_distribution": current_month_store_distribution,
        "top_products": top_products,
        "spending_trend": trend,
        "total_this_month": round(this_month_total, 2),
        "total_last_month": round(last_month_total, 2),
        "change_percent": round(change_percent, 1)
    }


@api_router.get("/backup/export")
async def export_data(device_id: str = Query(...)):
    receipts = await db.receipts.find({"device_id": device_id}, {"_id": 0}).to_list(10000)
    return {
        "device_id": device_id,
        "exported_at": datetime.now(timezone.utc).isoformat(),
        "total_receipts": len(receipts),
        "receipts": receipts
    }


# Admin endpoint to delete products by store name
@api_router.delete("/admin/products/by-store")
async def delete_products_by_store(store_name: str = Query(...)):
    """Delete all products from a specific store."""
    result = await db.products.delete_many({"store_name": {"$regex": store_name, "$options": "i"}})
    return {"deleted_count": result.deleted_count, "store_name": store_name}


# Admin endpoint to delete receipts by store name
@api_router.delete("/admin/receipts/by-store")
async def delete_receipts_by_store(store_name: str = Query(...), device_id: str = Query(...)):
    """Delete all receipts from a specific store for a device."""
    result = await db.receipts.delete_many({
        "device_id": device_id,
        "store_name": {"$regex": store_name, "$options": "i"}
    })
    return {"deleted_count": result.deleted_count, "store_name": store_name}


# VAT validation endpoint
@api_router.get("/stores/validate-vat")
async def validate_vat(vat: str = Query(...)):
    """Check if a VAT number is in our known stores list."""
    is_known = vat in STORE_VAT_MAPPING
    store_name = STORE_VAT_MAPPING.get(vat, None)
    return {
        "vat": vat,
        "is_known": is_known,
        "store_name": store_name
    }


# Get list of all supported stores
@api_router.get("/stores/supported")
async def get_supported_stores():
    """Get list of all supported stores with their VAT numbers."""
    stores = []
    seen = set()
    for vat, name in STORE_VAT_MAPPING.items():
        if name not in seen:
            stores.append({"name": name, "vat": vat})
            seen.add(name)
    return {"stores": sorted(stores, key=lambda x: x["name"])}


# Endpoint to request store review
@api_router.post("/stores/request-review")
async def request_store_review(
    vat: str = Body(...),
    store_name: str = Body(""),
    receipt_url: str = Body(""),
    device_id: str = Body(...)
):
    """Submit a store for review to be added to the supported list."""
    review_request = {
        "id": str(uuid.uuid4()),
        "vat": vat,
        "store_name": store_name,
        "receipt_url": receipt_url,
        "device_id": device_id,
        "status": "pending",
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.store_review_requests.insert_one(review_request)
    
    # Send email notification to admin (non-blocking)
    try:
        email_body = f"""
        <h2 style="color: #0D9488;">Νέα Αίτηση Καταστήματος</h2>
        <p>Υποβλήθηκε νέα αίτηση για προσθήκη καταστήματος:</p>
        <table style="width: 100%; border-collapse: collapse; margin: 20px 0;">
            <tr>
                <td style="padding: 10px; border-bottom: 1px solid #eee; font-weight: bold; width: 120px;">Κατάστημα:</td>
                <td style="padding: 10px; border-bottom: 1px solid #eee;">{store_name or 'Δεν δόθηκε'}</td>
            </tr>
            <tr>
                <td style="padding: 10px; border-bottom: 1px solid #eee; font-weight: bold;">ΑΦΜ:</td>
                <td style="padding: 10px; border-bottom: 1px solid #eee;">{vat}</td>
            </tr>
            <tr>
                <td style="padding: 10px; border-bottom: 1px solid #eee; font-weight: bold;">URL Απόδειξης:</td>
                <td style="padding: 10px; border-bottom: 1px solid #eee;">
                    <a href="{receipt_url}" style="color: #0D9488;">{receipt_url or 'Δεν δόθηκε'}</a>
                </td>
            </tr>
            <tr>
                <td style="padding: 10px; border-bottom: 1px solid #eee; font-weight: bold;">Device ID:</td>
                <td style="padding: 10px; border-bottom: 1px solid #eee; font-size: 12px; color: #666;">{device_id}</td>
            </tr>
        </table>
        <p style="margin-top: 20px;">
            <a href="/admin" style="display: inline-block; padding: 12px 24px; background: #0D9488; color: white; text-decoration: none; border-radius: 8px; font-weight: 500;">
                Διαχείριση Αιτήσεων
            </a>
        </p>
        """
        send_admin_notification(
            subject=f"🏪 Νέα Αίτηση Καταστήματος: {store_name or vat}",
            body=email_body
        )
    except Exception as e:
        logger.error(f"Failed to send notification email: {e}")
    
    return {"success": True, "message": "Request submitted for review", "request_id": review_request["id"]}


# ============ ADMIN ENDPOINTS ============

# Admin credentials - In production, use secure environment variables
ADMIN_USERNAME = os.environ.get("ADMIN_USERNAME", "admin")
ADMIN_PASSWORD = os.environ.get("ADMIN_PASSWORD", "admin2024!")

# ============ ADMIN AUTHENTICATION ============

class AdminLoginRequest(BaseModel):
    username: str
    password: str

@api_router.post("/admin/login")
async def admin_login(request: AdminLoginRequest):
    """Admin login with username/password."""
    if request.username != ADMIN_USERNAME or request.password != ADMIN_PASSWORD:
        raise HTTPException(status_code=401, detail="Invalid credentials")
    
    # Generate admin session token (simple JWT-like for admin)
    admin_token = jwt.encode(
        {
            "sub": "admin",
            "role": "admin",
            "exp": datetime.utcnow() + timedelta(hours=24),
            "iat": datetime.utcnow()
        },
        JWT_SECRET_KEY,
        algorithm=JWT_ALGORITHM
    )
    
    return {
        "success": True,
        "admin_token": admin_token,
        "expires_in": 86400  # 24 hours
    }

async def verify_admin_token(admin_token: str):
    """Verify admin token and return True if valid."""
    try:
        payload = jwt.decode(admin_token, JWT_SECRET_KEY, algorithms=[JWT_ALGORITHM])
        return payload.get("role") == "admin"
    except JWTError:
        return False

# ============ ADMIN RECEIPTS & PRODUCTS EXPORT ============

@api_router.get("/admin/receipts/all")
async def get_all_receipts(
    admin_key: str = Query(None),
    admin_token: str = Query(None),
    filter_store: str = Query(None),
    filter_user: str = Query(None),
    skip: int = 0,
    limit: int = 100
):
    """Get all receipts with filters (Admin only)."""
    # Verify admin access
    is_valid = admin_key == ADMIN_PASSWORD
    if not is_valid and admin_token:
        is_valid = await verify_admin_token(admin_token)
    if not is_valid:
        raise HTTPException(status_code=403, detail="Unauthorized")
    
    query = {}
    if filter_store:
        query["store_name"] = {"$regex": filter_store, "$options": "i"}
    if filter_user:
        query["device_id"] = filter_user
    
    receipts_cursor = await db.receipts.find(query).sort("created_at", -1).skip(skip).limit(limit).to_list(limit)
    total = await db.receipts.count_documents(query)
    
    # Build user lookup map (device_id -> user info)
    device_ids = list(set(r.get("device_id") for r in receipts_cursor if r.get("device_id")))
    users = await db.users.find({"device_id": {"$in": device_ids}}).to_list(1000)
    user_map = {}
    for u in users:
        if u.get("device_id"):
            user_map[u["device_id"]] = {
                "email": u.get("email", ""),
                "phone": u.get("phone", ""),
                "name": u.get("name", ""),
                "auth_provider": u.get("auth_provider", "unknown")
            }
    
    # Convert ObjectId to string and add user info
    receipts = []
    for r in receipts_cursor:
        if "_id" in r:
            r["_id"] = str(r["_id"])
        # Add user info
        device_id = r.get("device_id", "")
        user_info = user_map.get(device_id, {})
        r["user_email"] = user_info.get("email", "")
        r["user_phone"] = user_info.get("phone", "")
        r["user_name"] = user_info.get("name", "")
        r["auth_provider"] = user_info.get("auth_provider", "")
        receipts.append(r)
    
    return {
        "receipts": receipts,
        "total": total,
        "skip": skip,
        "limit": limit
    }

@api_router.get("/admin/receipts/export-excel")
async def export_all_receipts_excel(
    admin_key: str = Query(None),
    admin_token: str = Query(None),
    filter_store: str = Query(None),
    filter_user: str = Query(None)
):
    """Export all receipts and products to Excel (Admin only)."""
    # Verify admin access
    is_valid = admin_key == ADMIN_PASSWORD
    if not is_valid and admin_token:
        is_valid = await verify_admin_token(admin_token)
    if not is_valid:
        raise HTTPException(status_code=403, detail="Unauthorized")
    
    query = {}
    if filter_store:
        query["store_name"] = {"$regex": filter_store, "$options": "i"}
    if filter_user:
        query["device_id"] = filter_user
    
    receipts = await db.receipts.find(query).sort("created_at", -1).to_list(10000)
    
    # Create Excel workbook with multiple sheets
    wb = Workbook()
    
    # Sheet 1: Receipts Overview
    ws_receipts = wb.active
    ws_receipts.title = "Αποδείξεις"
    
    # Headers for receipts - UPDATED with more user details
    receipt_headers = ["ID Απόδειξης", "Device ID", "Email Χρήστη", "Τηλέφωνο", "Τρόπος Σύνδεσης", "Κατάστημα", "ΑΦΜ Καταστήματος", "Ημερομηνία", "Σύνολο €", "Τρόπος Πληρωμής", "Αρ. Προϊόντων"]
    
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="0D9488", end_color="0D9488", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    for col, header in enumerate(receipt_headers, 1):
        cell = ws_receipts.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center')
    
    # Get full user info for device_ids
    device_user_map = {}
    users = await db.users.find({}, {"device_id": 1, "email": 1, "phone": 1, "auth_provider": 1, "name": 1}).to_list(10000)
    for user in users:
        if user.get("device_id"):
            device_user_map[user["device_id"]] = {
                "email": user.get("email", ""),
                "phone": user.get("phone", ""),
                "auth_provider": user.get("auth_provider", "unknown"),
                "name": user.get("name", "")
            }
    
    # Auth provider display names
    auth_provider_names = {
        "email": "Email/Password",
        "google": "Google",
        "facebook": "Facebook",
        "apple": "Apple",
        "phone": "Phone SMS",
        "unknown": "Άγνωστο"
    }
    
    for row, receipt in enumerate(receipts, 2):
        device_id = receipt.get("device_id", "")
        user_info = device_user_map.get(device_id, {})
        
        ws_receipts.cell(row=row, column=1, value=str(receipt.get("id", receipt.get("_id", ""))))
        ws_receipts.cell(row=row, column=2, value=device_id)
        ws_receipts.cell(row=row, column=3, value=user_info.get("email", "Ανώνυμος"))
        ws_receipts.cell(row=row, column=4, value=user_info.get("phone", "-"))
        ws_receipts.cell(row=row, column=5, value=auth_provider_names.get(user_info.get("auth_provider", "unknown"), "Άγνωστο"))
        ws_receipts.cell(row=row, column=6, value=receipt.get("store_name", ""))
        ws_receipts.cell(row=row, column=7, value=receipt.get("store_vat", ""))
        ws_receipts.cell(row=row, column=8, value=receipt.get("date", ""))
        ws_receipts.cell(row=row, column=9, value=receipt.get("total", 0))
        ws_receipts.cell(row=row, column=10, value=receipt.get("payment_method", ""))
        ws_receipts.cell(row=row, column=11, value=len(receipt.get("items", [])))
    
    # Adjust column widths for receipts sheet
    col_widths_receipts = [22, 18, 30, 15, 18, 20, 15, 12, 12, 15, 12]
    for col, width in enumerate(col_widths_receipts, 1):
        ws_receipts.column_dimensions[get_column_letter(col)].width = width
    
    # Sheet 2: All Products
    ws_products = wb.create_sheet("Προϊόντα")
    
    product_headers = ["ID Απόδειξης", "Κατάστημα", "Ημερομηνία", "Κωδικός Προϊόντος", "Περιγραφή", "Ποσότητα", "Τιμή Μονάδας €", "Σύνολο €", "ΦΠΑ %"]
    
    for col, header in enumerate(product_headers, 1):
        cell = ws_products.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center')
    
    product_row = 2
    for receipt in receipts:
        receipt_id = receipt.get("id", receipt.get("_id", ""))
        store_name = receipt.get("store_name", "")
        date = receipt.get("date", "")
        
        for item in receipt.get("items", []):
            ws_products.cell(row=product_row, column=1, value=receipt_id)
            ws_products.cell(row=product_row, column=2, value=store_name)
            ws_products.cell(row=product_row, column=3, value=date)
            ws_products.cell(row=product_row, column=4, value=item.get("code", ""))
            ws_products.cell(row=product_row, column=5, value=item.get("description", ""))
            ws_products.cell(row=product_row, column=6, value=item.get("quantity", 1))
            ws_products.cell(row=product_row, column=7, value=item.get("unit_price", 0))
            ws_products.cell(row=product_row, column=8, value=item.get("total_value", 0))
            ws_products.cell(row=product_row, column=9, value=item.get("vat_percent", 0))
            product_row += 1
    
    # Adjust column widths for products sheet
    col_widths = [20, 20, 12, 15, 40, 10, 15, 12, 10]
    for col, width in enumerate(col_widths, 1):
        ws_products.column_dimensions[get_column_letter(col)].width = width
    
    # Sheet 3: Product Codes by Store
    ws_codes = wb.create_sheet("Κωδικοί_ανά_Κατάστημα")
    
    codes_headers = ["Κατάστημα", "Κωδικός Προϊόντος", "Περιγραφή", "Πλήθος Αγορών", "Μέση Τιμή €"]
    
    for col, header in enumerate(codes_headers, 1):
        cell = ws_codes.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center')
    
    # Aggregate product codes by store
    pipeline = [
        {"$unwind": "$items"},
        {"$match": {"items.code": {"$ne": ""}}},
        {"$group": {
            "_id": {"store": "$store_name", "code": "$items.code"},
            "description": {"$first": "$items.description"},
            "count": {"$sum": 1},
            "avg_price": {"$avg": "$items.unit_price"}
        }},
        {"$sort": {"_id.store": 1, "count": -1}}
    ]
    product_codes = await db.receipts.aggregate(pipeline).to_list(50000)
    
    for row, pc in enumerate(product_codes, 2):
        ws_codes.cell(row=row, column=1, value=pc["_id"].get("store", ""))
        ws_codes.cell(row=row, column=2, value=pc["_id"].get("code", ""))
        ws_codes.cell(row=row, column=3, value=pc.get("description", ""))
        ws_codes.cell(row=row, column=4, value=pc.get("count", 0))
        ws_codes.cell(row=row, column=5, value=round(pc.get("avg_price", 0), 2))
    
    # Adjust column widths for codes sheet
    codes_widths = [20, 15, 40, 15, 15]
    for col, width in enumerate(codes_widths, 1):
        ws_codes.column_dimensions[get_column_letter(col)].width = width
    
    # Save to BytesIO
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f"apodixxi_export_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

@api_router.get("/admin/users/all")
async def get_all_users_detailed(
    admin_key: str = Query(None),
    admin_token: str = Query(None),
    skip: int = 0,
    limit: int = 100
):
    """Get all users with receipt counts (Admin only)."""
    # Verify admin access
    is_valid = admin_key == ADMIN_PASSWORD
    if not is_valid and admin_token:
        is_valid = await verify_admin_token(admin_token)
    if not is_valid:
        raise HTTPException(status_code=403, detail="Unauthorized")
    
    users = await db.users.find().sort("created_at", -1).skip(skip).limit(limit).to_list(limit)
    total = await db.users.count_documents({})
    
    # Add receipt counts for each user
    for user in users:
        device_id = user.get("device_id")
        if device_id:
            receipt_count = await db.receipts.count_documents({"device_id": device_id})
            user["receipt_count"] = receipt_count
        else:
            user["receipt_count"] = 0
    
    return {
        "users": users,
        "total": total,
        "skip": skip,
        "limit": limit
    }

@api_router.get("/admin/products/existing")
async def get_existing_products(
    admin_key: str = Query(None),
    admin_token: str = Query(None),
    search: str = Query(None),
    store: str = Query(None),
    skip: int = 0,
    limit: int = 50
):
    """Get existing products from receipts for promotion selection (Admin only)."""
    # Verify admin access
    is_valid = admin_key == ADMIN_PASSWORD
    if not is_valid and admin_token:
        is_valid = await verify_admin_token(admin_token)
    if not is_valid:
        raise HTTPException(status_code=403, detail="Unauthorized")
    
    match_stage = {}
    if store:
        match_stage["store_name"] = {"$regex": store, "$options": "i"}
    
    pipeline = [
        {"$match": match_stage} if match_stage else {"$match": {}},
        {"$unwind": "$items"},
        {"$group": {
            "_id": {"$toUpper": "$items.description"},
            "code": {"$first": "$items.code"},
            "avg_price": {"$avg": "$items.unit_price"},
            "min_price": {"$min": "$items.unit_price"},
            "max_price": {"$max": "$items.unit_price"},
            "purchase_count": {"$sum": 1},
            "stores": {"$addToSet": "$store_name"}
        }},
        {"$sort": {"purchase_count": -1}}
    ]
    
    if search:
        pipeline.insert(1, {"$match": {"items.description": {"$regex": search, "$options": "i"}}})
    
    pipeline.append({"$skip": skip})
    pipeline.append({"$limit": limit})
    
    products = await db.receipts.aggregate(pipeline).to_list(limit)
    
    return {
        "products": [
            {
                "name": p["_id"],
                "code": p.get("code", ""),
                "avg_price": round(p.get("avg_price", 0), 2),
                "min_price": round(p.get("min_price", 0), 2),
                "max_price": round(p.get("max_price", 0), 2),
                "purchase_count": p.get("purchase_count", 0),
                "stores": p.get("stores", [])
            }
            for p in products
        ]
    }

# ============ ADMIN WEB DASHBOARD ============

@api_router.get("/admin/dashboard", response_class=HTMLResponse)
async def admin_dashboard():
    """Serve Admin Dashboard HTML."""
    html_content = '''<!DOCTYPE html>
<html lang="el">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>apodixxi Admin Dashboard</title>
    <link href="https://fonts.googleapis.com/css2?family=Inter:wght@400;500;600;700&display=swap" rel="stylesheet">
    <link href="https://cdn.jsdelivr.net/npm/lucide-static@latest/font/lucide.min.css" rel="stylesheet">
    <style>
        * { margin: 0; padding: 0; box-sizing: border-box; }
        body {
            font-family: 'Inter', -apple-system, BlinkMacSystemFont, sans-serif;
            background: #f5f7fa;
            min-height: 100vh;
        }
        
        /* Login Screen */
        .login-container {
            display: flex;
            align-items: center;
            justify-content: center;
            min-height: 100vh;
            background: linear-gradient(135deg, #0D9488 0%, #065F57 100%);
        }
        .login-box {
            background: white;
            padding: 40px;
            border-radius: 16px;
            box-shadow: 0 20px 60px rgba(0,0,0,0.3);
            width: 100%;
            max-width: 400px;
        }
        .login-logo {
            text-align: center;
            margin-bottom: 30px;
        }
        .login-logo h1 {
            color: #0D9488;
            font-size: 32px;
            display: flex;
            align-items: center;
            justify-content: center;
            gap: 10px;
        }
        .login-logo p {
            color: #666;
            margin-top: 5px;
        }
        .form-group {
            margin-bottom: 20px;
        }
        .form-group label {
            display: block;
            font-weight: 500;
            color: #333;
            margin-bottom: 8px;
        }
        .form-group input {
            width: 100%;
            padding: 14px 16px;
            border: 2px solid #e5e7eb;
            border-radius: 10px;
            font-size: 16px;
            transition: border-color 0.2s;
        }
        .form-group input:focus {
            outline: none;
            border-color: #0D9488;
        }
        .login-btn {
            width: 100%;
            padding: 14px;
            background: #0D9488;
            color: white;
            border: none;
            border-radius: 10px;
            font-size: 16px;
            font-weight: 600;
            cursor: pointer;
            transition: background 0.2s;
        }
        .login-btn:hover {
            background: #0b7d73;
        }
        .login-error {
            background: #fee2e2;
            color: #dc2626;
            padding: 12px;
            border-radius: 8px;
            margin-bottom: 20px;
            display: none;
        }
        
        /* Dashboard Layout */
        .dashboard { display: none; }
        .dashboard.active { display: block; }
        
        .sidebar {
            position: fixed;
            left: 0;
            top: 0;
            width: 260px;
            height: 100vh;
            background: linear-gradient(180deg, #0D9488 0%, #065F57 100%);
            padding: 20px;
            color: white;
            z-index: 100;
        }
        .sidebar-logo {
            padding: 10px 0 30px;
            border-bottom: 1px solid rgba(255,255,255,0.2);
            margin-bottom: 20px;
        }
        .sidebar-logo h1 {
            font-size: 24px;
            display: flex;
            align-items: center;
            gap: 10px;
        }
        .sidebar-logo p {
            font-size: 12px;
            opacity: 0.7;
            margin-top: 4px;
        }
        .nav-item {
            display: flex;
            align-items: center;
            gap: 12px;
            padding: 14px 16px;
            border-radius: 10px;
            cursor: pointer;
            transition: background 0.2s;
            margin-bottom: 6px;
        }
        .nav-item:hover, .nav-item.active {
            background: rgba(255,255,255,0.15);
        }
        .nav-item svg {
            width: 20px;
            height: 20px;
        }
        .logout-btn {
            position: absolute;
            bottom: 20px;
            left: 20px;
            right: 20px;
            padding: 12px;
            background: rgba(255,255,255,0.1);
            border: 1px solid rgba(255,255,255,0.2);
            border-radius: 10px;
            color: white;
            cursor: pointer;
            display: flex;
            align-items: center;
            justify-content: center;
            gap: 8px;
            transition: background 0.2s;
        }
        .logout-btn:hover {
            background: rgba(255,255,255,0.2);
        }
        
        .main-content {
            margin-left: 260px;
            padding: 30px;
            min-height: 100vh;
        }
        .page-header {
            margin-bottom: 30px;
        }
        .page-header h2 {
            font-size: 28px;
            color: #1f2937;
        }
        .page-header p {
            color: #6b7280;
            margin-top: 4px;
        }
        
        /* Stats Cards */
        .stats-grid {
            display: grid;
            grid-template-columns: repeat(auto-fit, minmax(220px, 1fr));
            gap: 20px;
            margin-bottom: 30px;
        }
        .stat-card {
            background: white;
            padding: 24px;
            border-radius: 12px;
            box-shadow: 0 2px 8px rgba(0,0,0,0.06);
        }
        .stat-card .label {
            color: #6b7280;
            font-size: 14px;
            margin-bottom: 8px;
        }
        .stat-card .value {
            font-size: 32px;
            font-weight: 700;
            color: #1f2937;
        }
        .stat-card.highlight {
            background: linear-gradient(135deg, #0D9488 0%, #065F57 100%);
        }
        .stat-card.highlight .label,
        .stat-card.highlight .value {
            color: white;
        }
        
        /* Tables */
        .card {
            background: white;
            border-radius: 12px;
            box-shadow: 0 2px 8px rgba(0,0,0,0.06);
            overflow: hidden;
            margin-bottom: 24px;
        }
        .card-header {
            display: flex;
            justify-content: space-between;
            align-items: center;
            padding: 20px 24px;
            border-bottom: 1px solid #e5e7eb;
        }
        .card-header h3 {
            font-size: 18px;
            color: #1f2937;
        }
        .card-body {
            padding: 0;
        }
        table {
            width: 100%;
            border-collapse: collapse;
        }
        th, td {
            padding: 14px 20px;
            text-align: left;
        }
        th {
            background: #f9fafb;
            font-weight: 600;
            color: #374151;
            font-size: 13px;
            text-transform: uppercase;
        }
        td {
            border-top: 1px solid #e5e7eb;
            color: #4b5563;
        }
        tr:hover td {
            background: #f9fafb;
        }
        
        /* Buttons */
        .btn {
            padding: 10px 20px;
            border-radius: 8px;
            font-weight: 500;
            cursor: pointer;
            border: none;
            transition: all 0.2s;
            display: inline-flex;
            align-items: center;
            gap: 8px;
        }
        .btn-primary {
            background: #0D9488;
            color: white;
        }
        .btn-primary:hover {
            background: #0b7d73;
        }
        .btn-secondary {
            background: #e5e7eb;
            color: #374151;
        }
        .btn-secondary:hover {
            background: #d1d5db;
        }
        .btn-success {
            background: #059669;
            color: white;
        }
        .btn-danger {
            background: #dc2626;
            color: white;
        }
        
        /* Filters */
        .filters {
            display: flex;
            gap: 12px;
            margin-bottom: 20px;
            flex-wrap: wrap;
        }
        .filter-input {
            padding: 10px 16px;
            border: 2px solid #e5e7eb;
            border-radius: 8px;
            font-size: 14px;
            min-width: 200px;
        }
        .filter-input:focus {
            outline: none;
            border-color: #0D9488;
        }
        
        /* Modal */
        .modal-overlay {
            position: fixed;
            top: 0;
            left: 0;
            right: 0;
            bottom: 0;
            background: rgba(0,0,0,0.5);
            display: flex;
            align-items: center;
            justify-content: center;
            z-index: 1000;
            display: none;
        }
        .modal-overlay.active {
            display: flex;
        }
        .modal {
            background: white;
            border-radius: 16px;
            width: 100%;
            max-width: 600px;
            max-height: 90vh;
            overflow-y: auto;
        }
        .modal-header {
            padding: 20px 24px;
            border-bottom: 1px solid #e5e7eb;
            display: flex;
            justify-content: space-between;
            align-items: center;
        }
        .modal-header h3 {
            font-size: 20px;
            color: #1f2937;
        }
        .modal-close {
            background: none;
            border: none;
            font-size: 24px;
            cursor: pointer;
            color: #6b7280;
        }
        .modal-body {
            padding: 24px;
        }
        .modal-footer {
            padding: 16px 24px;
            border-top: 1px solid #e5e7eb;
            display: flex;
            justify-content: flex-end;
            gap: 12px;
        }
        
        /* Page sections */
        .page { display: none; }
        .page.active { display: block; }
        
        /* Badge */
        .badge {
            display: inline-block;
            padding: 4px 10px;
            border-radius: 20px;
            font-size: 12px;
            font-weight: 500;
        }
        .badge-success {
            background: #d1fae5;
            color: #065f46;
        }
        .badge-warning {
            background: #fef3c7;
            color: #92400e;
        }
        .badge-info {
            background: #dbeafe;
            color: #1e40af;
        }
        
        /* Loading */
        .loading {
            display: flex;
            justify-content: center;
            align-items: center;
            padding: 40px;
        }
        .spinner {
            width: 40px;
            height: 40px;
            border: 3px solid #e5e7eb;
            border-top-color: #0D9488;
            border-radius: 50%;
            animation: spin 1s linear infinite;
        }
        @keyframes spin {
            to { transform: rotate(360deg); }
        }
        
        /* Responsive */
        @media (max-width: 1024px) {
            .sidebar { width: 80px; padding: 10px; }
            .sidebar-logo h1 span, .nav-item span, .logout-btn span { display: none; }
            .main-content { margin-left: 80px; }
        }
    </style>
</head>
<body>
    <!-- Login Screen -->
    <div class="login-container" id="loginScreen">
        <div class="login-box">
            <div class="login-logo">
                <h1>🧾 apodixxi</h1>
                <p>Admin Dashboard</p>
            </div>
            <div class="login-error" id="loginError"></div>
            <form id="loginForm">
                <div class="form-group">
                    <label>Username</label>
                    <input type="text" id="username" placeholder="admin" required>
                </div>
                <div class="form-group">
                    <label>Password</label>
                    <input type="password" id="password" placeholder="••••••••" required>
                </div>
                <button type="submit" class="login-btn">Σύνδεση</button>
            </form>
        </div>
    </div>
    
    <!-- Dashboard -->
    <div class="dashboard" id="dashboard">
        <aside class="sidebar">
            <div class="sidebar-logo">
                <h1>🧾 <span>apodixxi</span></h1>
                <p>Admin Panel</p>
            </div>
            <nav>
                <div class="nav-item active" data-page="overview">
                    <svg viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2"><rect x="3" y="3" width="7" height="7"/><rect x="14" y="3" width="7" height="7"/><rect x="14" y="14" width="7" height="7"/><rect x="3" y="14" width="7" height="7"/></svg>
                    <span>Επισκόπηση</span>
                </div>
                <div class="nav-item" data-page="receipts">
                    <svg viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2"><path d="M14 2H6a2 2 0 0 0-2 2v16a2 2 0 0 0 2 2h12a2 2 0 0 0 2-2V8z"/><polyline points="14 2 14 8 20 8"/><line x1="16" y1="13" x2="8" y2="13"/><line x1="16" y1="17" x2="8" y2="17"/></svg>
                    <span>Αποδείξεις</span>
                </div>
                <div class="nav-item" data-page="users">
                    <svg viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2"><path d="M17 21v-2a4 4 0 0 0-4-4H5a4 4 0 0 0-4 4v2"/><circle cx="9" cy="7" r="4"/><path d="M23 21v-2a4 4 0 0 0-3-3.87"/><path d="M16 3.13a4 4 0 0 1 0 7.75"/></svg>
                    <span>Χρήστες</span>
                </div>
                <div class="nav-item" data-page="promotions">
                    <svg viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2"><polygon points="12 2 15.09 8.26 22 9.27 17 14.14 18.18 21.02 12 17.77 5.82 21.02 7 14.14 2 9.27 8.91 8.26 12 2"/></svg>
                    <span>Προωθήσεις</span>
                </div>
                <div class="nav-item" data-page="promo-codes">
                    <svg viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2"><path d="M21 4H3a2 2 0 0 0-2 2v12a2 2 0 0 0 2 2h18a2 2 0 0 0 2-2V6a2 2 0 0 0-2-2z"/><path d="M7 12h10"/></svg>
                    <span>Promo Codes</span>
                </div>
            </nav>
            <button class="logout-btn" onclick="logout()">
                <svg viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2" width="20" height="20"><path d="M9 21H5a2 2 0 0 1-2-2V5a2 2 0 0 1 2-2h4"/><polyline points="16 17 21 12 16 7"/><line x1="21" y1="12" x2="9" y2="12"/></svg>
                <span>Αποσύνδεση</span>
            </button>
        </aside>
        
        <main class="main-content">
            <!-- Overview Page -->
            <div class="page active" id="page-overview">
                <div class="page-header">
                    <h2>Επισκόπηση</h2>
                    <p>Συνολικά στατιστικά της εφαρμογής</p>
                </div>
                <div class="stats-grid" id="statsGrid">
                    <div class="loading"><div class="spinner"></div></div>
                </div>
                <div class="card">
                    <div class="card-header">
                        <h3>Κορυφαία Καταστήματα</h3>
                    </div>
                    <div class="card-body">
                        <table>
                            <thead><tr><th>Κατάστημα</th><th>Αποδείξεις</th></tr></thead>
                            <tbody id="topStoresTable"></tbody>
                        </table>
                    </div>
                </div>
            </div>
            
            <!-- Receipts Page -->
            <div class="page" id="page-receipts">
                <div class="page-header">
                    <h2>Αποδείξεις</h2>
                    <p>Διαχείριση και εξαγωγή αποδείξεων</p>
                </div>
                <div class="filters">
                    <input type="text" class="filter-input" id="filterStore" placeholder="Φίλτρο κατά κατάστημα...">
                    <input type="text" class="filter-input" id="filterUser" placeholder="Φίλτρο κατά device_id...">
                    <button class="btn btn-primary" onclick="loadReceipts()">Αναζήτηση</button>
                    <button class="btn btn-success" onclick="exportExcel()">📊 Εξαγωγή Excel</button>
                </div>
                <div class="card">
                    <div class="card-body">
                        <table>
                            <thead><tr><th>ID</th><th>Email</th><th>Τηλέφωνο</th><th>Τρόπος Σύνδεσης</th><th>Κατάστημα</th><th>Ημερομηνία</th><th>Σύνολο</th><th>Προϊόντα</th></tr></thead>
                            <tbody id="receiptsTable"></tbody>
                        </table>
                    </div>
                </div>
                <div id="receiptsPagination" style="padding: 20px; text-align: center;"></div>
            </div>
            
            <!-- Users Page -->
            <div class="page" id="page-users">
                <div class="page-header">
                    <h2>Χρήστες</h2>
                    <p>Διαχείριση χρηστών</p>
                </div>
                <div class="card">
                    <div class="card-body">
                        <table>
                            <thead><tr><th>Email</th><th>Όνομα</th><th>Τύπος</th><th>Αποδείξεις</th><th>Εγγραφή</th><th>Ενέργειες</th></tr></thead>
                            <tbody id="usersTable"></tbody>
                        </table>
                    </div>
                </div>
            </div>
            
            <!-- Promotions Page -->
            <div class="page" id="page-promotions">
                <div class="page-header">
                    <h2>Προωθήσεις Προϊόντων</h2>
                    <p>Διαχείριση promoted προϊόντων στις προτάσεις</p>
                </div>
                <div class="filters">
                    <button class="btn btn-primary" onclick="openPromotionModal()">+ Νέα Προώθηση</button>
                    <button class="btn btn-secondary" onclick="openExistingProductsModal()">Επιλογή από Υπάρχοντα</button>
                </div>
                <div class="card">
                    <div class="card-body">
                        <table>
                            <thead><tr><th>Τίτλος</th><th>Προϊόν</th><th>Τιμή</th><th>Κατάστημα</th><th>Κατάσταση</th><th>Views</th><th>Clicks</th><th>Ενέργειες</th></tr></thead>
                            <tbody id="promotionsTable"></tbody>
                        </table>
                    </div>
                </div>
            </div>
            
            <!-- Promo Codes Page -->
            <div class="page" id="page-promo-codes">
                <div class="page-header">
                    <h2>Promo Codes</h2>
                    <p>Διαχείριση κωδικών προσφορών για Premium</p>
                </div>
                <div class="filters">
                    <button class="btn btn-primary" onclick="openPromoCodeModal()">+ Νέος Κωδικός</button>
                </div>
                <div class="card">
                    <div class="card-body">
                        <table>
                            <thead><tr><th>Κωδικός</th><th>Διάρκεια</th><th>Χρήσεις</th><th>Κατάσταση</th><th>Ενέργειες</th></tr></thead>
                            <tbody id="promoCodesTable"></tbody>
                        </table>
                    </div>
                </div>
            </div>
        </main>
    </div>
    
    <!-- Promotion Modal -->
    <div class="modal-overlay" id="promotionModal">
        <div class="modal">
            <div class="modal-header">
                <h3>Νέα Προώθηση Προϊόντος</h3>
                <button class="modal-close" onclick="closeModal('promotionModal')">&times;</button>
            </div>
            <div class="modal-body">
                <form id="promotionForm">
                    <div class="form-group">
                        <label>Τίτλος Προώθησης *</label>
                        <input type="text" id="promoTitle" required placeholder="π.χ. Μεγάλη Προσφορά!">
                    </div>
                    <div class="form-group">
                        <label>Περιγραφή</label>
                        <input type="text" id="promoDescription" placeholder="Περιγραφή προσφοράς...">
                    </div>
                    <div class="form-group">
                        <label>Όνομα Προϊόντος *</label>
                        <input type="text" id="promoProductName" required placeholder="π.χ. ΓΑΛΑ ΦΑΡΜΑ 1L">
                    </div>
                    <div style="display: grid; grid-template-columns: 1fr 1fr; gap: 16px;">
                        <div class="form-group">
                            <label>Τιμή Προσφοράς (€)</label>
                            <input type="number" step="0.01" id="promoPrice" placeholder="1.99">
                        </div>
                        <div class="form-group">
                            <label>Αρχική Τιμή (€)</label>
                            <input type="number" step="0.01" id="promoOriginalPrice" placeholder="2.49">
                        </div>
                    </div>
                    <div class="form-group">
                        <label>Κατάστημα</label>
                        <input type="text" id="promoStoreName" placeholder="π.χ. ΣΚΛΑΒΕΝΙΤΗΣ">
                    </div>
                    <div class="form-group">
                        <label>Κωδικός Barcode</label>
                        <input type="text" id="promoBarcodeCode" placeholder="5201234567890">
                    </div>
                    <div class="form-group">
                        <label>URL Εικόνας</label>
                        <input type="text" id="promoImageUrl" placeholder="https://...">
                    </div>
                    <div style="display: grid; grid-template-columns: 1fr 1fr; gap: 16px;">
                        <div class="form-group">
                            <label>Ημ/νία Έναρξης</label>
                            <input type="date" id="promoStartDate">
                        </div>
                        <div class="form-group">
                            <label>Ημ/νία Λήξης</label>
                            <input type="date" id="promoEndDate">
                        </div>
                    </div>
                    <div class="form-group">
                        <label>Προτεραιότητα (0-100)</label>
                        <input type="number" id="promoPriority" value="10" min="0" max="100">
                    </div>
                </form>
            </div>
            <div class="modal-footer">
                <button class="btn btn-secondary" onclick="closeModal('promotionModal')">Ακύρωση</button>
                <button class="btn btn-primary" onclick="savePromotion()">Αποθήκευση</button>
            </div>
        </div>
    </div>
    
    <!-- Existing Products Modal -->
    <div class="modal-overlay" id="existingProductsModal">
        <div class="modal" style="max-width: 800px;">
            <div class="modal-header">
                <h3>Επιλογή από Υπάρχοντα Προϊόντα</h3>
                <button class="modal-close" onclick="closeModal('existingProductsModal')">&times;</button>
            </div>
            <div class="modal-body">
                <div class="filters">
                    <input type="text" class="filter-input" id="searchExistingProduct" placeholder="Αναζήτηση προϊόντος...">
                    <button class="btn btn-primary" onclick="searchExistingProducts()">Αναζήτηση</button>
                </div>
                <table>
                    <thead><tr><th>Προϊόν</th><th>Κωδικός</th><th>Μ.Ο. Τιμή</th><th>Αγορές</th><th></th></tr></thead>
                    <tbody id="existingProductsTable"></tbody>
                </table>
            </div>
        </div>
    </div>
    
    <!-- Promo Code Modal -->
    <div class="modal-overlay" id="promoCodeModal">
        <div class="modal">
            <div class="modal-header">
                <h3>Νέος Promo Code</h3>
                <button class="modal-close" onclick="closeModal('promoCodeModal')">&times;</button>
            </div>
            <div class="modal-body">
                <form id="promoCodeForm">
                    <div class="form-group">
                        <label>Κωδικός *</label>
                        <input type="text" id="pcCode" required placeholder="π.χ. SUMMER2025" style="text-transform: uppercase;">
                    </div>
                    <div class="form-group">
                        <label>Διάρκεια Premium (ημέρες)</label>
                        <input type="number" id="pcDays" value="30" min="1">
                    </div>
                    <div class="form-group">
                        <label>Μέγιστες Χρήσεις (κενό = απεριόριστο)</label>
                        <input type="number" id="pcMaxUses" placeholder="100">
                    </div>
                </form>
            </div>
            <div class="modal-footer">
                <button class="btn btn-secondary" onclick="closeModal('promoCodeModal')">Ακύρωση</button>
                <button class="btn btn-primary" onclick="savePromoCode()">Δημιουργία</button>
            </div>
        </div>
    </div>

    <script>
        const API_BASE = '/api';
        let adminToken = localStorage.getItem('adminToken');
        let currentReceiptsPage = 0;
        
        // Check if already logged in
        if (adminToken) {
            showDashboard();
            loadOverview();
        }
        
        // Login
        document.getElementById('loginForm').addEventListener('submit', async (e) => {
            e.preventDefault();
            const username = document.getElementById('username').value;
            const password = document.getElementById('password').value;
            
            try {
                const res = await fetch(`${API_BASE}/admin/login`, {
                    method: 'POST',
                    headers: { 'Content-Type': 'application/json' },
                    body: JSON.stringify({ username, password })
                });
                
                const data = await res.json();
                
                if (res.ok) {
                    adminToken = data.admin_token;
                    localStorage.setItem('adminToken', adminToken);
                    showDashboard();
                    loadOverview();
                } else {
                    showLoginError(data.detail || 'Σφάλμα σύνδεσης');
                }
            } catch (err) {
                showLoginError('Σφάλμα σύνδεσης');
            }
        });
        
        function showLoginError(msg) {
            const el = document.getElementById('loginError');
            el.textContent = msg;
            el.style.display = 'block';
        }
        
        function showDashboard() {
            document.getElementById('loginScreen').style.display = 'none';
            document.getElementById('dashboard').classList.add('active');
        }
        
        function logout() {
            localStorage.removeItem('adminToken');
            location.reload();
        }
        
        // Navigation
        document.querySelectorAll('.nav-item').forEach(item => {
            item.addEventListener('click', () => {
                const page = item.dataset.page;
                
                document.querySelectorAll('.nav-item').forEach(i => i.classList.remove('active'));
                item.classList.add('active');
                
                document.querySelectorAll('.page').forEach(p => p.classList.remove('active'));
                document.getElementById('page-' + page).classList.add('active');
                
                // Load page data
                if (page === 'overview') loadOverview();
                else if (page === 'receipts') loadReceipts();
                else if (page === 'users') loadUsers();
                else if (page === 'promotions') loadPromotions();
                else if (page === 'promo-codes') loadPromoCodes();
            });
        });
        
        // API calls
        async function apiCall(endpoint, options = {}) {
            const url = new URL(API_BASE + endpoint, window.location.origin);
            url.searchParams.append('admin_token', adminToken);
            
            const res = await fetch(url, options);
            if (res.status === 403) {
                logout();
                return null;
            }
            return res;
        }
        
        // Format auth provider for display
        function formatAuthProvider(provider) {
            const providers = {
                'email': '📧 Email',
                'google': '🔵 Google',
                'facebook': '🔷 Facebook',
                'apple': '🍎 Apple',
                'phone': '📱 Phone',
                'unknown': '❓ Unknown'
            };
            return providers[provider] || providers['unknown'];
        }
        
        async function loadOverview() {
            try {
                const res = await apiCall('/admin/stats');
                const data = await res.json();
                
                document.getElementById('statsGrid').innerHTML = `
                    <div class="stat-card highlight">
                        <div class="label">Χρήστες</div>
                        <div class="value">${data.total_users || 0}</div>
                    </div>
                    <div class="stat-card">
                        <div class="label">Premium</div>
                        <div class="value">${data.paid_users || 0}</div>
                    </div>
                    <div class="stat-card">
                        <div class="label">Αποδείξεις</div>
                        <div class="value">${data.total_receipts || 0}</div>
                    </div>
                    <div class="stat-card">
                        <div class="label">Προϊόντα</div>
                        <div class="value">${data.total_products || 0}</div>
                    </div>
                    <div class="stat-card">
                        <div class="label">Promo Codes</div>
                        <div class="value">${data.total_promo_codes || 0}</div>
                    </div>
                `;
                
                const topStores = data.top_stores || [];
                document.getElementById('topStoresTable').innerHTML = topStores.map(s => `
                    <tr>
                        <td>${s.name || 'Άγνωστο'}</td>
                        <td><strong>${s.count}</strong></td>
                    </tr>
                `).join('');
            } catch (err) {
                console.error('Error loading overview:', err);
            }
        }
        
        async function loadReceipts() {
            const filterStore = document.getElementById('filterStore').value;
            const filterUser = document.getElementById('filterUser').value;
            
            let endpoint = `/admin/receipts/all?skip=${currentReceiptsPage * 50}&limit=50`;
            if (filterStore) endpoint += `&filter_store=${encodeURIComponent(filterStore)}`;
            if (filterUser) endpoint += `&filter_user=${encodeURIComponent(filterUser)}`;
            
            try {
                const res = await apiCall(endpoint);
                const data = await res.json();
                
                document.getElementById('receiptsTable').innerHTML = data.receipts.map(r => `
                    <tr>
                        <td title="${r.id || r._id}">${(r.id || r._id || '').substring(0, 8)}...</td>
                        <td>${r.user_email || '-'}</td>
                        <td>${r.user_phone || '-'}</td>
                        <td><span class="badge badge-info">${formatAuthProvider(r.auth_provider)}</span></td>
                        <td>${r.store_name || '-'}</td>
                        <td>${r.date || '-'}</td>
                        <td><strong>€${(r.total || 0).toFixed(2)}</strong></td>
                        <td>${(r.items || []).length}</td>
                    </tr>
                `).join('');
                
                const totalPages = Math.ceil(data.total / 50);
                document.getElementById('receiptsPagination').innerHTML = `
                    Σελίδα ${currentReceiptsPage + 1} από ${totalPages} (${data.total} αποδείξεις)
                    ${currentReceiptsPage > 0 ? '<button class="btn btn-secondary" onclick="prevReceiptsPage()">◀ Προηγούμενη</button>' : ''}
                    ${currentReceiptsPage < totalPages - 1 ? '<button class="btn btn-secondary" onclick="nextReceiptsPage()">Επόμενη ▶</button>' : ''}
                `;
            } catch (err) {
                console.error('Error loading receipts:', err);
            }
        }
        
        function prevReceiptsPage() { currentReceiptsPage--; loadReceipts(); }
        function nextReceiptsPage() { currentReceiptsPage++; loadReceipts(); }
        
        async function exportExcel() {
            const filterStore = document.getElementById('filterStore').value;
            const filterUser = document.getElementById('filterUser').value;
            
            let url = `${API_BASE}/admin/receipts/export-excel?admin_token=${adminToken}`;
            if (filterStore) url += `&filter_store=${encodeURIComponent(filterStore)}`;
            if (filterUser) url += `&filter_user=${encodeURIComponent(filterUser)}`;
            
            window.open(url, '_blank');
        }
        
        async function loadUsers() {
            try {
                const res = await apiCall('/admin/users/all?limit=100');
                const data = await res.json();
                
                document.getElementById('usersTable').innerHTML = data.users.map(u => `
                    <tr>
                        <td>${u.email || '-'}</td>
                        <td>${u.name || '-'}</td>
                        <td><span class="badge ${u.account_type === 'paid' ? 'badge-success' : 'badge-info'}">${u.account_type === 'paid' ? 'Premium' : 'Free'}</span></td>
                        <td>${u.receipt_count || 0}</td>
                        <td>${u.created_at ? new Date(u.created_at).toLocaleDateString('el-GR') : '-'}</td>
                        <td>
                            <button class="btn btn-secondary" onclick="upgradeUser('${u._id}')">Αναβάθμιση</button>
                        </td>
                    </tr>
                `).join('');
            } catch (err) {
                console.error('Error loading users:', err);
            }
        }
        
        async function upgradeUser(userId) {
            const days = prompt('Πόσες ημέρες Premium;', '30');
            if (!days) return;
            
            try {
                await fetch(`${API_BASE}/admin/users/${userId}/upgrade?admin_token=${adminToken}`, {
                    method: 'POST',
                    headers: { 'Content-Type': 'application/json' },
                    body: JSON.stringify({ days: parseInt(days) })
                });
                loadUsers();
                alert('Ο χρήστης αναβαθμίστηκε!');
            } catch (err) {
                alert('Σφάλμα');
            }
        }
        
        async function loadPromotions() {
            try {
                const res = await apiCall('/admin/promotions');
                const data = await res.json();
                
                document.getElementById('promotionsTable').innerHTML = (data.promotions || []).map(p => `
                    <tr>
                        <td>${p.title || '-'}</td>
                        <td>${p.product_name || '-'}</td>
                        <td>${p.price ? '€' + p.price.toFixed(2) : '-'} ${p.original_price ? '<small style="text-decoration:line-through;color:#999">€' + p.original_price.toFixed(2) + '</small>' : ''}</td>
                        <td>${p.store_name || '-'}</td>
                        <td><span class="badge ${p.is_active ? 'badge-success' : 'badge-warning'}">${p.is_active ? 'Ενεργή' : 'Ανενεργή'}</span></td>
                        <td>${p.views_count || 0}</td>
                        <td>${p.clicks_count || 0}</td>
                        <td>
                            <button class="btn btn-secondary" onclick="togglePromotion('${p._id}', ${!p.is_active})">${p.is_active ? 'Απενεργ.' : 'Ενεργ.'}</button>
                            <button class="btn btn-danger" onclick="deletePromotion('${p._id}')">Διαγραφή</button>
                        </td>
                    </tr>
                `).join('') || '<tr><td colspan="8" style="text-align:center;padding:40px;color:#999;">Δεν υπάρχουν προωθήσεις</td></tr>';
            } catch (err) {
                console.error('Error loading promotions:', err);
            }
        }
        
        async function loadPromoCodes() {
            try {
                const res = await apiCall('/admin/promo-codes');
                const data = await res.json();
                
                document.getElementById('promoCodesTable').innerHTML = (data.promo_codes || []).map(p => `
                    <tr>
                        <td><strong>${p.code}</strong></td>
                        <td>${p.duration_days} ημέρες</td>
                        <td>${p.used_count || 0}${p.max_uses ? '/' + p.max_uses : ''}</td>
                        <td><span class="badge ${p.is_active ? 'badge-success' : 'badge-warning'}">${p.is_active ? 'Ενεργός' : 'Ανενεργός'}</span></td>
                        <td>
                            <button class="btn btn-secondary" onclick="togglePromoCode('${p.code}', ${!p.is_active})">${p.is_active ? 'Απενεργ.' : 'Ενεργ.'}</button>
                            <button class="btn btn-danger" onclick="deletePromoCode('${p.code}')">Διαγραφή</button>
                        </td>
                    </tr>
                `).join('') || '<tr><td colspan="5" style="text-align:center;padding:40px;color:#999;">Δεν υπάρχουν promo codes</td></tr>';
            } catch (err) {
                console.error('Error loading promo codes:', err);
            }
        }
        
        // Modals
        function openPromotionModal() {
            document.getElementById('promotionForm').reset();
            document.getElementById('promotionModal').classList.add('active');
        }
        
        function openPromoCodeModal() {
            document.getElementById('promoCodeForm').reset();
            document.getElementById('promoCodeModal').classList.add('active');
        }
        
        function closeModal(id) {
            document.getElementById(id).classList.remove('active');
        }
        
        async function savePromotion() {
            const promo = {
                title: document.getElementById('promoTitle').value,
                description: document.getElementById('promoDescription').value,
                product_name: document.getElementById('promoProductName').value,
                price: parseFloat(document.getElementById('promoPrice').value) || null,
                original_price: parseFloat(document.getElementById('promoOriginalPrice').value) || null,
                store_name: document.getElementById('promoStoreName').value,
                barcode_code: document.getElementById('promoBarcodeCode').value,
                image_url: document.getElementById('promoImageUrl').value,
                start_date: document.getElementById('promoStartDate').value || null,
                end_date: document.getElementById('promoEndDate').value || null,
                priority: parseInt(document.getElementById('promoPriority').value) || 10,
                target_all_users: true
            };
            
            try {
                await fetch(`${API_BASE}/admin/promotions?admin_token=${adminToken}`, {
                    method: 'POST',
                    headers: { 'Content-Type': 'application/json' },
                    body: JSON.stringify(promo)
                });
                closeModal('promotionModal');
                loadPromotions();
                alert('Η προώθηση δημιουργήθηκε!');
            } catch (err) {
                alert('Σφάλμα');
            }
        }
        
        async function togglePromotion(id, isActive) {
            try {
                await fetch(`${API_BASE}/admin/promotions/${id}?admin_token=${adminToken}`, {
                    method: 'PATCH',
                    headers: { 'Content-Type': 'application/json' },
                    body: JSON.stringify({ is_active: isActive })
                });
                loadPromotions();
            } catch (err) {
                alert('Σφάλμα');
            }
        }
        
        async function deletePromotion(id) {
            if (!confirm('Διαγραφή προώθησης;')) return;
            try {
                await fetch(`${API_BASE}/admin/promotions/${id}?admin_token=${adminToken}`, { method: 'DELETE' });
                loadPromotions();
            } catch (err) {
                alert('Σφάλμα');
            }
        }
        
        async function savePromoCode() {
            const code = document.getElementById('pcCode').value.toUpperCase();
            const days = parseInt(document.getElementById('pcDays').value);
            const maxUses = document.getElementById('pcMaxUses').value ? parseInt(document.getElementById('pcMaxUses').value) : null;
            
            try {
                await fetch(`${API_BASE}/admin/promo-codes?admin_token=${adminToken}`, {
                    method: 'POST',
                    headers: { 'Content-Type': 'application/json' },
                    body: JSON.stringify({ code, duration_days: days, max_uses: maxUses })
                });
                closeModal('promoCodeModal');
                loadPromoCodes();
                alert('Ο κωδικός δημιουργήθηκε!');
            } catch (err) {
                alert('Σφάλμα');
            }
        }
        
        async function togglePromoCode(code, isActive) {
            try {
                await fetch(`${API_BASE}/admin/promo-codes/${code}/toggle?is_active=${isActive}&admin_token=${adminToken}`, { method: 'PATCH' });
                loadPromoCodes();
            } catch (err) {
                alert('Σφάλμα');
            }
        }
        
        async function deletePromoCode(code) {
            if (!confirm('Διαγραφή κωδικού ' + code + ';')) return;
            try {
                await fetch(`${API_BASE}/admin/promo-codes/${code}?admin_token=${adminToken}`, { method: 'DELETE' });
                loadPromoCodes();
            } catch (err) {
                alert('Σφάλμα');
            }
        }
        
        // Existing products modal
        async function openExistingProductsModal() {
            document.getElementById('existingProductsModal').classList.add('active');
            searchExistingProducts();
        }
        
        async function searchExistingProducts() {
            const search = document.getElementById('searchExistingProduct').value;
            let endpoint = `/admin/products/existing?limit=20`;
            if (search) endpoint += `&search=${encodeURIComponent(search)}`;
            
            try {
                const res = await apiCall(endpoint);
                const data = await res.json();
                
                document.getElementById('existingProductsTable').innerHTML = (data.products || []).map(p => `
                    <tr>
                        <td>${p.name}</td>
                        <td>${p.code || '-'}</td>
                        <td>€${p.avg_price.toFixed(2)}</td>
                        <td>${p.purchase_count}</td>
                        <td><button class="btn btn-primary" onclick="selectExistingProduct('${p.name.replace(/'/g, "\\'")}', ${p.avg_price})">Επιλογή</button></td>
                    </tr>
                `).join('') || '<tr><td colspan="5" style="text-align:center;padding:20px;">Δεν βρέθηκαν προϊόντα</td></tr>';
            } catch (err) {
                console.error('Error:', err);
            }
        }
        
        function selectExistingProduct(name, price) {
            closeModal('existingProductsModal');
            document.getElementById('promoProductName').value = name;
            document.getElementById('promoOriginalPrice').value = price.toFixed(2);
            openPromotionModal();
        }
    </script>
</body>
</html>'''
    return html_content

@api_router.get("/admin/store-reviews")
async def get_store_reviews(
    admin_key: str = Query(...),
    status: str = Query(default="pending")
):
    """Get all store review requests (Admin only)."""
    if admin_key != ADMIN_PASSWORD:
        raise HTTPException(status_code=403, detail="Unauthorized")
    
    query = {} if status == "all" else {"status": status}
    reviews = await db.store_review_requests.find(query, {"_id": 0}).sort("created_at", -1).to_list(100)
    
    return {
        "total": len(reviews),
        "reviews": reviews
    }


@api_router.post("/admin/store-reviews/{review_id}/approve")
async def approve_store_review(
    review_id: str,
    admin_key: str = Query(...)
):
    """Approve a store review request and add it to known stores (Admin only)."""
    if admin_key != ADMIN_PASSWORD:
        raise HTTPException(status_code=403, detail="Unauthorized")
    
    review = await db.store_review_requests.find_one({"id": review_id})
    if not review:
        raise HTTPException(status_code=404, detail="Review not found")
    
    # Update the review status
    await db.store_review_requests.update_one(
        {"id": review_id},
        {"$set": {
            "status": "approved",
            "approved_at": datetime.now(timezone.utc).isoformat()
        }}
    )
    
    # Note: To actually add to STORE_VAT_MAPPING, you need to update the code
    # For now, we'll save it to a separate collection
    await db.approved_stores.update_one(
        {"vat": review["vat"]},
        {"$set": {
            "vat": review["vat"],
            "store_name": review["store_name"],
            "approved_at": datetime.now(timezone.utc).isoformat()
        }},
        upsert=True
    )
    
    return {"success": True, "message": f"Store {review['store_name']} approved"}


@api_router.post("/admin/store-reviews/{review_id}/reject")
async def reject_store_review(
    review_id: str,
    admin_key: str = Query(...),
    reason: str = Body(default="")
):
    """Reject a store review request (Admin only)."""
    if admin_key != ADMIN_PASSWORD:
        raise HTTPException(status_code=403, detail="Unauthorized")
    
    review = await db.store_review_requests.find_one({"id": review_id})
    if not review:
        raise HTTPException(status_code=404, detail="Review not found")
    
    await db.store_review_requests.update_one(
        {"id": review_id},
        {"$set": {
            "status": "rejected",
            "rejected_at": datetime.now(timezone.utc).isoformat(),
            "rejection_reason": reason
        }}
    )
    
    return {"success": True, "message": "Review rejected"}


@api_router.delete("/admin/store-reviews/{review_id}")
async def delete_store_review(
    review_id: str,
    admin_key: str = Query(...)
):
    """Delete a store review request (Admin only)."""
    if admin_key != ADMIN_PASSWORD:
        raise HTTPException(status_code=403, detail="Unauthorized")
    
    result = await db.store_review_requests.delete_one({"id": review_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Review not found")
    
    return {"success": True, "message": "Review deleted"}


@api_router.get("/admin/stats")
async def get_admin_stats(
    admin_key: str = Query(None),
    admin_token: str = Query(None)
):
    """Get admin statistics (Admin only)."""
    # Verify admin access
    is_valid = admin_key == ADMIN_PASSWORD
    if not is_valid and admin_token:
        is_valid = await verify_admin_token(admin_token)
    if not is_valid:
        raise HTTPException(status_code=403, detail="Unauthorized")
    
    total_devices = await db.devices.count_documents({})
    total_receipts = await db.receipts.count_documents({})
    total_products = await db.products.count_documents({})
    pending_reviews = await db.store_review_requests.count_documents({"status": "pending"})
    approved_stores = await db.approved_stores.count_documents({})
    total_users = await db.users.count_documents({})
    total_promo_codes = await db.promo_codes.count_documents({})
    paid_users = await db.users.count_documents({"account_type": "paid"})
    
    # Get top stores by receipt count
    pipeline = [
        {"$group": {"_id": "$store_name", "count": {"$sum": 1}}},
        {"$sort": {"count": -1}},
        {"$limit": 10}
    ]
    top_stores = await db.receipts.aggregate(pipeline).to_list(10)
    
    return {
        "total_devices": total_devices,
        "total_receipts": total_receipts,
        "total_products": total_products,
        "pending_reviews": pending_reviews,
        "approved_stores": approved_stores,
        "total_users": total_users,
        "total_promo_codes": total_promo_codes,
        "paid_users": paid_users,
        "top_stores": [{"name": s["_id"], "count": s["count"]} for s in top_stores]
    }


# ============ EMAIL NOTIFICATIONS ============

# Email settings from environment
ADMIN_EMAIL = os.environ.get("ADMIN_EMAIL", "admin@example.com")
SMTP_HOST = os.environ.get("SMTP_HOST", "smtp.gmail.com")
SMTP_PORT = int(os.environ.get("SMTP_PORT", "587"))
SMTP_USER = os.environ.get("SMTP_USER", "")
SMTP_PASSWORD = os.environ.get("SMTP_PASSWORD", "")
SMTP_FROM = os.environ.get("SMTP_FROM", "noreply@grocerytracker.app")

def send_admin_notification(subject: str, body: str):
    """Send email notification to admin. Non-blocking, logs errors."""
    if not SMTP_USER or not SMTP_PASSWORD:
        logger.info(f"Email notification skipped (SMTP not configured): {subject}")
        return False
    
    try:
        msg = MIMEMultipart('alternative')
        msg['Subject'] = subject
        msg['From'] = SMTP_FROM
        msg['To'] = ADMIN_EMAIL
        
        # HTML email body
        html_body = f"""
        <html>
        <body style="font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', Roboto, sans-serif; background: #f5f5f5; padding: 20px;">
            <div style="max-width: 600px; margin: 0 auto; background: white; border-radius: 12px; padding: 24px; box-shadow: 0 2px 8px rgba(0,0,0,0.1);">
                <div style="text-align: center; margin-bottom: 20px;">
                    <h1 style="color: #0D9488; margin: 0;">🧾 apodixxi</h1>
                    <p style="color: #666; margin: 5px 0 0;">Admin Notification</p>
                </div>
                <hr style="border: none; border-top: 1px solid #eee; margin: 20px 0;">
                {body}
                <hr style="border: none; border-top: 1px solid #eee; margin: 20px 0;">
                <p style="color: #999; font-size: 12px; text-align: center;">
                    Αυτό το email στάλθηκε αυτόματα από το apodixxi Admin System.
                </p>
            </div>
        </body>
        </html>
        """
        
        msg.attach(MIMEText(body, 'plain'))
        msg.attach(MIMEText(html_body, 'html'))
        
        with smtplib.SMTP(SMTP_HOST, SMTP_PORT) as server:
            server.starttls()
            server.login(SMTP_USER, SMTP_PASSWORD)
            server.send_message(msg)
        
        logger.info(f"Email notification sent: {subject}")
        return True
    except Exception as e:
        logger.error(f"Failed to send email: {e}")
        return False


# ============ AI INTEGRATION WITH GEMINI ============

from emergentintegrations.llm.chat import LlmChat, UserMessage

EMERGENT_LLM_KEY = os.environ.get("EMERGENT_LLM_KEY", "")

class AIInsightRequest(BaseModel):
    device_id: str
    insight_type: str = "spending"  # spending, savings, recommendations

class AIChatRequest(BaseModel):
    device_id: str
    message: str
    session_id: Optional[str] = None

class AIRecommendationRequest(BaseModel):
    device_id: str
    category: Optional[str] = None
    limit: int = 5

@api_router.post("/ai/insights")
async def get_ai_insights(request: AIInsightRequest):
    """Get AI-powered insights about user's shopping habits."""
    if not EMERGENT_LLM_KEY:
        raise HTTPException(status_code=500, detail="AI service not configured")
    
    # Fetch user's receipt data
    receipts = await db.receipts.find({"device_id": request.device_id}).to_list(100)
    
    if not receipts:
        return {"insight": "Δεν υπάρχουν αρκετά δεδομένα για ανάλυση. Σκανάρετε περισσότερες αποδείξεις!"}
    
    # Calculate statistics
    total_spent = sum(r.get("total_amount", r.get("total", 0)) for r in receipts)
    stores = {}
    categories = {}
    products = {}
    
    for receipt in receipts:
        store = receipt.get("store_name", "Άγνωστο")
        stores[store] = stores.get(store, 0) + receipt.get("total_amount", receipt.get("total", 0))
        
        for item in receipt.get("items", []):
            cat = item.get("category", "Άλλο")
            categories[cat] = categories.get(cat, 0) + item.get("total_price", 0)
            
            name = item.get("name", "")
            if name:
                products[name] = products.get(name, 0) + item.get("quantity", 1)
    
    # Prepare context for AI
    top_stores = sorted(stores.items(), key=lambda x: x[1], reverse=True)[:5]
    top_categories = sorted(categories.items(), key=lambda x: x[1], reverse=True)[:5]
    top_products = sorted(products.items(), key=lambda x: x[1], reverse=True)[:10]
    
    context = f"""
Στοιχεία χρήστη apodixxi:
- Συνολικές αποδείξεις: {len(receipts)}
- Συνολικά έξοδα: {total_spent:.2f}€
- Top καταστήματα: {', '.join([f'{s[0]} ({s[1]:.2f}€)' for s in top_stores])}
- Top κατηγορίες: {', '.join([f'{c[0]} ({c[1]:.2f}€)' for c in top_categories])}
- Συχνά προϊόντα: {', '.join([f'{p[0]} (x{p[1]})' for p in top_products])}
"""
    
    # Create AI prompt based on insight type
    if request.insight_type == "spending":
        prompt = f"""Με βάση τα παρακάτω δεδομένα αγορών, δώσε μια σύντομη ανάλυση (2-3 προτάσεις) για τα έξοδα του χρήστη στα ελληνικά:
{context}

Δώσε πρακτικές συμβουλές για εξοικονόμηση χρημάτων."""
    
    elif request.insight_type == "savings":
        prompt = f"""Με βάση τα παρακάτω δεδομένα αγορών, προτείνε 3 συγκεκριμένους τρόπους εξοικονόμησης στα ελληνικά:
{context}

Να είσαι συγκεκριμένος με ποσά και καταστήματα."""
    
    else:  # recommendations
        prompt = f"""Με βάση τα παρακάτω δεδομένα αγορών, προτείνε προϊόντα ή προσφορές που μπορεί να ενδιαφέρουν τον χρήστη στα ελληνικά:
{context}

Προτείνε 3-5 προϊόντα με βάση τις αγοραστικές του συνήθειες."""
    
    try:
        chat = LlmChat(
            api_key=EMERGENT_LLM_KEY,
            session_id=f"insights_{request.device_id}_{request.insight_type}",
            system_message="Είσαι ο AI βοηθός του apodixxi, μιας εφαρμογής παρακολούθησης αποδείξεων. Δίνεις σύντομες, πρακτικές συμβουλές στα ελληνικά."
        ).with_model("gemini", "gemini-2.5-flash")
        
        user_message = UserMessage(text=prompt)
        response = await chat.send_message(user_message)
        
        return {
            "insight": response,
            "stats": {
                "total_receipts": len(receipts),
                "total_spent": round(total_spent, 2),
                "top_store": top_stores[0][0] if top_stores else None,
                "top_category": top_categories[0][0] if top_categories else None
            }
        }
    except Exception as e:
        logger.error(f"AI insight error: {e}")
        raise HTTPException(status_code=500, detail=f"AI error: {str(e)}")


@api_router.post("/ai/chat")
async def ai_chat(request: AIChatRequest):
    """Chat with AI assistant about shopping habits."""
    if not EMERGENT_LLM_KEY:
        raise HTTPException(status_code=500, detail="AI service not configured")
    
    # Fetch user's receipt data for context
    receipts = await db.receipts.find({"device_id": request.device_id}).to_list(50)
    
    # Build context
    total_spent = sum(r.get("total_amount", r.get("total", 0)) for r in receipts)
    recent_receipts = receipts[:5] if receipts else []
    
    context = f"""
Πληροφορίες χρήστη:
- Αριθμός αποδείξεων: {len(receipts)}
- Συνολικά έξοδα: {total_spent:.2f}€
- Πρόσφατες αγορές: {', '.join([r.get('store_name', 'Άγνωστο') for r in recent_receipts])}
"""
    
    session_id = request.session_id or f"chat_{request.device_id}_{uuid.uuid4().hex[:8]}"
    
    try:
        chat = LlmChat(
            api_key=EMERGENT_LLM_KEY,
            session_id=session_id,
            system_message=f"""Είσαι ο AI βοηθός του apodixxi, μιας εφαρμογής παρακολούθησης αποδείξεων supermarket στην Ελλάδα.
Απαντάς πάντα στα ελληνικά, σύντομα και φιλικά.
{context}
Μπορείς να βοηθήσεις με ερωτήσεις σχετικά με:
- Έξοδα και στατιστικά αγορών
- Συμβουλές εξοικονόμησης
- Σύγκριση τιμών μεταξύ καταστημάτων
- Προτάσεις προϊόντων"""
        ).with_model("gemini", "gemini-2.5-flash")
        
        user_message = UserMessage(text=request.message)
        response = await chat.send_message(user_message)
        
        # Store chat message in database
        await db.ai_chats.insert_one({
            "device_id": request.device_id,
            "session_id": session_id,
            "user_message": request.message,
            "ai_response": response,
            "created_at": datetime.now(timezone.utc).isoformat()
        })
        
        return {
            "response": response,
            "session_id": session_id
        }
    except Exception as e:
        logger.error(f"AI chat error: {e}")
        raise HTTPException(status_code=500, detail=f"AI error: {str(e)}")


@api_router.post("/ai/smart-recommendations")
async def get_ai_recommendations(request: AIRecommendationRequest):
    """Get AI-powered product recommendations based on shopping history."""
    if not EMERGENT_LLM_KEY:
        raise HTTPException(status_code=500, detail="AI service not configured")
    
    # Fetch user's purchase history
    receipts = await db.receipts.find({"device_id": request.device_id}).to_list(100)
    
    if not receipts:
        return {"recommendations": [], "message": "Σκανάρετε αποδείξεις για εξατομικευμένες προτάσεις!"}
    
    # Analyze purchase patterns
    products = {}
    stores = set()
    categories = {}
    
    for receipt in receipts:
        stores.add(receipt.get("store_name", ""))
        for item in receipt.get("items", []):
            name = item.get("name", "")
            cat = item.get("category", "Άλλο")
            price = item.get("unit_price", 0)
            
            if name:
                if name not in products:
                    products[name] = {"count": 0, "total_spent": 0, "category": cat, "avg_price": 0}
                products[name]["count"] += item.get("quantity", 1)
                products[name]["total_spent"] += item.get("total_price", 0)
    
    top_products = sorted(products.items(), key=lambda x: x[1]["count"], reverse=True)[:15]
    
    category_filter = f"για την κατηγορία {request.category}" if request.category else ""
    
    prompt = f"""Με βάση το ιστορικό αγορών του χρήστη, προτείνε {request.limit} προϊόντα {category_filter} στα ελληνικά.

Συχνά αγοραζόμενα προϊόντα:
{chr(10).join([f'- {p[0]} (x{p[1]["count"]}, κατηγορία: {p[1]["category"]})' for p in top_products])}

Καταστήματα που ψωνίζει: {', '.join(list(stores)[:5])}

Απάντησε σε μορφή JSON array με objects που έχουν: title, description, reason, estimated_savings (προαιρετικό)
Παράδειγμα: [{{"title": "Γάλα σε προσφορά", "description": "Αγοράζετε συχνά γάλα - δείτε προσφορές", "reason": "Βασισμένο στις αγορές σας"}}]"""

    try:
        chat = LlmChat(
            api_key=EMERGENT_LLM_KEY,
            session_id=f"recs_{request.device_id}",
            system_message="Είσαι ένας έξυπνος βοηθός αγορών. Απαντάς ΜΟΝΟ με valid JSON, χωρίς markdown."
        ).with_model("gemini", "gemini-2.5-flash")
        
        user_message = UserMessage(text=prompt)
        response = await chat.send_message(user_message)
        
        # Try to parse JSON response
        import json
        try:
            # Clean up response if it has markdown
            clean_response = response.strip()
            if clean_response.startswith("```"):
                clean_response = clean_response.split("```")[1]
                if clean_response.startswith("json"):
                    clean_response = clean_response[4:]
            recommendations = json.loads(clean_response)
        except:
            # If JSON parsing fails, return as text
            recommendations = [{"title": "AI Προτάσεις", "description": response, "reason": "AI-generated"}]
        
        return {
            "recommendations": recommendations[:request.limit],
            "source": "ai",
            "based_on_receipts": len(receipts)
        }
    except Exception as e:
        logger.error(f"AI recommendations error: {e}")
        raise HTTPException(status_code=500, detail=f"AI error: {str(e)}")


@api_router.get("/ai/weekly-summary")
async def get_weekly_summary(device_id: str):
    """Get AI-generated weekly shopping summary."""
    if not EMERGENT_LLM_KEY:
        raise HTTPException(status_code=500, detail="AI service not configured")
    
    # Get receipts from last 7 days
    week_ago = datetime.now(timezone.utc) - timedelta(days=7)
    
    receipts = await db.receipts.find({
        "device_id": device_id,
        "created_at": {"$gte": week_ago.isoformat()}
    }).to_list(100)
    
    if not receipts:
        # Try getting all receipts if none in last week
        receipts = await db.receipts.find({"device_id": device_id}).to_list(20)
    
    if not receipts:
        return {
            "summary": "Δεν υπάρχουν αποδείξεις αυτή την εβδομάδα. Σκανάρετε τις αποδείξεις σας!",
            "stats": {}
        }
    
    total = sum(r.get("total_amount", r.get("total", 0)) for r in receipts)
    stores = {}
    items_count = 0
    
    for r in receipts:
        store = r.get("store_name", "Άγνωστο")
        stores[store] = stores.get(store, 0) + r.get("total_amount", r.get("total", 0))
        items_count += len(r.get("items", []))
    
    prompt = f"""Δημιούργησε μια σύντομη εβδομαδιαία ανασκόπηση αγορών (3-4 προτάσεις) στα ελληνικά:

- Συνολικά έξοδα: {total:.2f}€
- Αριθμός αποδείξεων: {len(receipts)}
- Αριθμός προϊόντων: {items_count}
- Καταστήματα: {', '.join([f'{s} ({v:.2f}€)' for s, v in stores.items()])}

Συμπεριέλαβε:
1. Σύνοψη εξόδων
2. Μια παρατήρηση για τις αγορές
3. Μια συμβουλή εξοικονόμησης"""

    try:
        chat = LlmChat(
            api_key=EMERGENT_LLM_KEY,
            session_id=f"weekly_{device_id}",
            system_message="Είσαι φιλικός οικονομικός σύμβουλος. Δίνεις σύντομες, χρήσιμες συμβουλές στα ελληνικά."
        ).with_model("gemini", "gemini-2.5-flash")
        
        user_message = UserMessage(text=prompt)
        response = await chat.send_message(user_message)
        
        return {
            "summary": response,
            "stats": {
                "total_spent": round(total, 2),
                "receipts_count": len(receipts),
                "items_count": items_count,
                "stores": stores
            }
        }
    except Exception as e:
        logger.error(f"Weekly summary error: {e}")
        raise HTTPException(status_code=500, detail=f"AI error: {str(e)}")


# ============ IN-APP PURCHASES ============

class PurchaseVerifyRequest(BaseModel):
    receipt: str
    product_id: str
    platform: str  # 'ios' or 'android'
    user_email: Optional[str] = None

class PurchaseHistoryItem(BaseModel):
    product_id: str
    purchase_date: str
    expires_at: Optional[str] = None
    status: str

@api_router.post("/purchases/verify")
async def verify_purchase(request: PurchaseVerifyRequest):
    """Verify and process an in-app purchase."""
    
    # In production, you would verify the receipt with Apple/Google
    # For now, we trust the purchase and upgrade the user
    
    # Determine subscription duration based on product
    if "monthly" in request.product_id.lower():
        duration_days = 30
        plan_name = "apodixxi+ Μηνιαία"
    elif "yearly" in request.product_id.lower():
        duration_days = 365
        plan_name = "apodixxi+ Ετήσια"
    else:
        duration_days = 30
        plan_name = "apodixxi+"
    
    expires_at = datetime.now(timezone.utc) + timedelta(days=duration_days)
    
    # Update user subscription
    if request.user_email:
        user = await db.users.find_one({"email": request.user_email.lower()})
        if user:
            await db.users.update_one(
                {"_id": user["_id"]},
                {"$set": {
                    "account_type": "paid",
                    "subscription_expires_at": expires_at.isoformat(),
                    "subscription_plan": plan_name,
                    "last_purchase_date": datetime.now(timezone.utc).isoformat()
                }}
            )
    
    # Store purchase record
    purchase_record = {
        "id": str(uuid.uuid4()),
        "user_email": request.user_email,
        "product_id": request.product_id,
        "platform": request.platform,
        "receipt": request.receipt[:100] + "..." if len(request.receipt) > 100 else request.receipt,  # Truncate for storage
        "purchase_date": datetime.now(timezone.utc).isoformat(),
        "expires_at": expires_at.isoformat(),
        "status": "active"
    }
    
    await db.purchases.insert_one(purchase_record)
    
    logger.info(f"Purchase verified: {request.product_id} for {request.user_email}")
    
    return {
        "success": True,
        "message": "Η αγορά επαληθεύτηκε με επιτυχία!",
        "subscription": {
            "plan": plan_name,
            "expires_at": expires_at.isoformat(),
            "status": "active"
        }
    }


@api_router.get("/purchases/status")
async def get_purchase_status(user_email: str):
    """Get user's subscription status."""
    user = await db.users.find_one({"email": user_email.lower()})
    
    if not user:
        return {"status": "free", "has_subscription": False}
    
    account_type = user.get("account_type", "free")
    expires_at = user.get("subscription_expires_at")
    
    # Check if subscription is expired
    if expires_at:
        expires_date = datetime.fromisoformat(expires_at.replace('Z', '+00:00'))
        if expires_date < datetime.now(timezone.utc):
            account_type = "expired"
    
    return {
        "status": account_type,
        "has_subscription": account_type == "paid",
        "subscription_plan": user.get("subscription_plan"),
        "expires_at": expires_at
    }


@api_router.post("/purchases/restore")
async def restore_purchases(user_email: str):
    """Restore purchases for a user."""
    # Find active purchases for this user
    purchases = await db.purchases.find({
        "user_email": user_email.lower(),
        "status": "active"
    }).to_list(10)
    
    if not purchases:
        return {"success": False, "message": "Δεν βρέθηκαν προηγούμενες αγορές"}
    
    # Find the most recent valid purchase
    valid_purchase = None
    for purchase in purchases:
        expires_at = purchase.get("expires_at")
        if expires_at:
            expires_date = datetime.fromisoformat(expires_at.replace('Z', '+00:00'))
            if expires_date > datetime.now(timezone.utc):
                valid_purchase = purchase
                break
    
    if valid_purchase:
        # Restore subscription
        await db.users.update_one(
            {"email": user_email.lower()},
            {"$set": {
                "account_type": "paid",
                "subscription_expires_at": valid_purchase["expires_at"],
                "subscription_plan": valid_purchase.get("product_id", "apodixxi+")
            }}
        )
        return {
            "success": True,
            "message": "Η συνδρομή σου επαναφέρθηκε!",
            "subscription": {
                "expires_at": valid_purchase["expires_at"]
            }
        }
    
    return {"success": False, "message": "Δεν βρέθηκαν ενεργές συνδρομές"}


app.include_router(api_router)

app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=["*"],
    allow_methods=["*"],
    allow_headers=["*"],
)

@app.on_event("shutdown")
async def shutdown_db_client():
    client.close()
