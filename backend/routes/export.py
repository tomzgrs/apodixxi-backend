"""Data export routes for paid users."""
import io
from datetime import datetime, timezone
from fastapi import APIRouter, HTTPException, Depends
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from ..config import db
from .auth import get_current_user

router = APIRouter(prefix="/export", tags=["Export"])

# ============ HELPER FUNCTIONS ============

def check_user_is_paid(user: dict) -> bool:
    """Check if user has active paid subscription."""
    if not user:
        return False
    
    account_type = user.get('account_type', 'free')
    if account_type == 'free':
        return False
    
    expires_at = user.get('subscription_expires_at')
    if expires_at:
        try:
            if isinstance(expires_at, str):
                expiry_dt = datetime.fromisoformat(expires_at.replace('Z', '+00:00'))
            else:
                expiry_dt = expires_at
            if expiry_dt < datetime.now(timezone.utc):
                return False
        except:
            pass
    
    return True

# ============ ENDPOINTS ============

@router.get("/check-access")
async def check_export_access(user: dict = Depends(get_current_user)):
    """Check if user has access to export features."""
    is_paid = check_user_is_paid(user)
    return {
        "has_access": is_paid,
        "account_type": user.get("account_type", "free"),
        "subscription_expires_at": user.get("subscription_expires_at")
    }

@router.get("/receipts")
async def export_receipts_excel(user: dict = Depends(get_current_user)):
    """Export all receipts to Excel file (paid users only)."""
    if not check_user_is_paid(user):
        raise HTTPException(
            status_code=403, 
            detail="Αυτή η λειτουργία είναι διαθέσιμη μόνο για apodixxi+ χρήστες"
        )
    
    device_id = user.get("device_id")
    if not device_id:
        raise HTTPException(status_code=400, detail="Device ID not found")
    
    receipts = await db.receipts.find({"device_id": device_id}).to_list(10000)
    
    if not receipts:
        raise HTTPException(status_code=404, detail="Δεν βρέθηκαν αποδείξεις")
    
    # Create workbook
    wb = Workbook()
    
    # ===== SUMMARY SHEET =====
    ws_summary = wb.active
    ws_summary.title = "Σύνοψη"
    
    # Styles
    header_font = Font(bold=True, size=12, color="FFFFFF")
    header_fill = PatternFill(start_color="0D9488", end_color="0D9488", fill_type="solid")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Calculate statistics
    total_spent = sum(r.get("total_amount", r.get("total", 0)) for r in receipts)
    stores = {}
    for r in receipts:
        store = r.get("store_name", "Άγνωστο")
        stores[store] = stores.get(store, 0) + r.get("total_amount", r.get("total", 0))
    
    # Write summary
    ws_summary['A1'] = 'apodixxi+ - Αναφορά Αγορών'
    ws_summary['A1'].font = Font(bold=True, size=16, color="0D9488")
    ws_summary.merge_cells('A1:D1')
    
    ws_summary['A3'] = 'Ημερομηνία Εξαγωγής:'
    ws_summary['B3'] = datetime.now().strftime('%d/%m/%Y %H:%M')
    
    ws_summary['A4'] = 'Σύνολο Αποδείξεων:'
    ws_summary['B4'] = len(receipts)
    
    ws_summary['A5'] = 'Συνολικά Έξοδα:'
    ws_summary['B5'] = f'{total_spent:.2f}€'
    
    ws_summary['A7'] = 'Έξοδα ανά Κατάστημα'
    ws_summary['A7'].font = Font(bold=True, size=12)
    
    row = 8
    for store, amount in sorted(stores.items(), key=lambda x: x[1], reverse=True):
        ws_summary[f'A{row}'] = store
        ws_summary[f'B{row}'] = f'{amount:.2f}€'
        row += 1
    
    # ===== RECEIPTS SHEET =====
    ws_receipts = wb.create_sheet("Αποδείξεις")
    
    headers = ['Ημερομηνία', 'Κατάστημα', 'Σύνολο', 'Αριθμός Προϊόντων', 'Τρόπος Πληρωμής']
    for col, header in enumerate(headers, 1):
        cell = ws_receipts.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal='center')
    
    for row_idx, receipt in enumerate(receipts, 2):
        ws_receipts.cell(row=row_idx, column=1, value=receipt.get('date', ''))
        ws_receipts.cell(row=row_idx, column=2, value=receipt.get('store_name', ''))
        ws_receipts.cell(row=row_idx, column=3, value=receipt.get('total_amount', receipt.get('total', 0)))
        ws_receipts.cell(row=row_idx, column=4, value=len(receipt.get('items', [])))
        ws_receipts.cell(row=row_idx, column=5, value=receipt.get('payment_method', ''))
    
    # ===== PRODUCTS SHEET =====
    ws_products = wb.create_sheet("Προϊόντα")
    
    product_headers = ['Προϊόν', 'Κατάστημα', 'Ποσότητα', 'Τιμή Μονάδας', 'Σύνολο', 'Ημερομηνία']
    for col, header in enumerate(product_headers, 1):
        cell = ws_products.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal='center')
    
    product_row = 2
    for receipt in receipts:
        for item in receipt.get('items', []):
            ws_products.cell(row=product_row, column=1, value=item.get('description', item.get('name', '')))
            ws_products.cell(row=product_row, column=2, value=receipt.get('store_name', ''))
            ws_products.cell(row=product_row, column=3, value=item.get('quantity', 1))
            ws_products.cell(row=product_row, column=4, value=item.get('unit_price', 0))
            ws_products.cell(row=product_row, column=5, value=item.get('total_value', item.get('total_price', 0)))
            ws_products.cell(row=product_row, column=6, value=receipt.get('date', ''))
            product_row += 1
    
    # Adjust column widths
    for ws in [ws_summary, ws_receipts, ws_products]:
        for column_cells in ws.columns:
            max_length = 0
            column = column_cells[0].column_letter
            for cell in column_cells:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column].width = adjusted_width
    
    # Save to bytes
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f"apodixxi_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )
